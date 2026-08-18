"""
Cockpit appels — Hympyr Énergies
=================================

Outil de pilotage de la campagne d'appels clients, en équipe.

TROIS PROFILS
-------------
  Chloé, Patricia  — commerciales : appels clients, points de livraison, export.
  Mika             — manager      : tableau de bord uniquement.

CE QUI PERMET DE TRAVAILLER À PLUSIEURS SANS SE MARCHER DESSUS
--------------------------------------------------------------
1. Chaque enregistrement porte le nom de qui l'a fait (colonne traite_par).
   Une pastille colorée l'affiche à côté du nom du client.
2. Un verrou de consultation signale en temps réel qu'une collègue est déjà sur
   la fiche, avec depuis combien de temps. Il n'interdit rien : il informe, ce
   qui suffit à éviter le double appel.
3. Des filtres « non traités / traités par Chloé / traités par Patricia »
   permettent à chacune de se réserver une portion de la file.
4. Le tableau de bord du manager est alimenté directement par l'activité des
   deux commerciales, sans ressaisie.

PERSISTANCE
-----------
La base SQLite est écrite dans DOSSIER_DONNEES. Sur une plateforme à système de
fichiers éphémère, ce disque est perdu à chaque redémarrage du conteneur : les
exports du soir restent alors la seule sauvegarde. Définir la variable
d'environnement HYMPYR_DATA_DIR vers un volume persistant pour fermer le sujet.

DONNÉES PERSONNELLES
--------------------
Cet outil traite noms, adresses, téléphones et e-mails de clients : information
classifiée C2. Hébergement et plateforme doivent être instruits en conséquence,
et l'application inscrite au registre des applications.

Dépendances : streamlit >= 1.31, pandas >= 2.0, openpyxl >= 3.1
"""

from __future__ import annotations

import hashlib
import io
import os
import random
import re
import sqlite3
import time
import datetime as dt
from contextlib import contextmanager
from pathlib import Path
from typing import Iterable

import pandas as pd
import streamlit as st

# ─────────────────────────────────────────────────────────────────────────────
# CONFIGURATION
# ─────────────────────────────────────────────────────────────────────────────

DOSSIER_DONNEES = Path(os.environ.get("HYMPYR_DATA_DIR", Path(__file__).parent))
DOSSIER_DONNEES.mkdir(parents=True, exist_ok=True)

DB_PATH = DOSSIER_DONNEES / "suivi_appels.db"
DOSSIER_SAUVEGARDES = DOSSIER_DONNEES / "sauvegardes"
DOSSIER_SAUVEGARDES.mkdir(parents=True, exist_ok=True)

VERT, VERT_FONCE, ORANGE, GRIS = "#1A6B45", "#0D3D27", "#FF5C29", "#8E9A94"
BLEU = "#2F5D8C"   # profil administrateur

# ── Comptes ──────────────────────────────────────────────────────────────────
# Les mots de passe ne figurent jamais en clair : seule leur empreinte SHA-256
# salée est stockée. Pour changer un mot de passe, recalculer l'empreinte avec
# empreinte("nouveau_mot_de_passe") et remplacer la valeur ci-dessous.
# En production, préférer st.secrets — voir charger_empreinte().
SEL = "hympyr-cockpit-2026"

PROFILS = {
    "Chloé": {
        "role": "commercial",
        "couleur": VERT,
        "empreinte": "3eb1540ff1bcc49460df5b6df9b6e650ec4b49c69b269c56f968123164aa3a3b",
    },
    "Patricia": {
        "role": "commercial",
        "couleur": ORANGE,
        "empreinte": "ddff2e48a6a2eb29792cec7c7bf4f7cac6694b737dffa4ec6a8e9b52a0547766",
    },
    "Mika": {
        "role": "manager",
        "couleur": VERT_FONCE,
        "empreinte": "e97cafd6d1baa1ede105c0811b8e26d144505b1cb49d9107f720e6d424202a47",
    },
    "Quentin": {
        "role": "admin",
        "couleur": BLEU,
        # Mot de passe provisoire : QDEadmin8131@!!  — à changer, voir empreinte().
        "empreinte": "c6d1892f279428bcd9e60975e74de88fef32ba28c6d080d5475640d60e5c10cc",
    },
}

LIBELLES_ROLES = {"commercial": "Commerciale", "manager": "Manager", "admin": "Administrateur"}

# Qui apparaît dans le suivi de performance du tableau de bord.
COMMERCIALES = [nom for nom, p in PROFILS.items() if p["role"] == "commercial"]
# Qui peut traiter une fiche, donc apparaître dans les filtres et les pastilles.
TRAITANTS = [nom for nom, p in PROFILS.items() if p["role"] in ("commercial", "admin")]
# Le RESET reste exclusif au manager.
ROLES_RESET = {"manager"}

PRODUITS = [
    "GNR", "Gasoil routier", "Sans plomb", "AdBlue",
    "Fioul domestique", "HVO", "Granulés de bois", "Lubrifiants / Huiles",
]

STATUTS = [
    "À appeler", "À rappeler", "Injoignable",
    "Fait ✅", "Doublon", "Ancien client (à sortir)",
]
STATUTS_TERMINES = {"Fait ✅", "Doublon", "Ancien client (à sortir)"}
STATUTS_ADRESSE = ["À vérifier", "Vérifié ✅", "Adresse obsolète"]

# Champs que l'on peut demander à vider explicitement, et leur libellé.
LIBELLES_CHAMPS = {
    "Produits achetés": "produits",
    "E-mail confirmé": "email_maj",
    "Téléphone confirmé": "tel_maj",
    "Doublon du client n°": "doublon_de",
    "Motif de sortie": "motif_sortie",
    "Notes": "note",
}

MOTIFS_SORTIE = [
    "—", "Passé à la concurrence", "Utilise une autre énergie",
    "Décès", "Cessation d'activité / fermeture", "Ne souhaite plus être contacté",
    "Injoignable définitivement", "Autre",
]

# Échéance : obligation d'émission de la facturation électronique pour les PME.
DEADLINE = dt.date(2027, 9, 1)

# Durée pendant laquelle un verrou de consultation reste considéré comme actif.
VERROU_MINUTES = 12

# Libellés d'affichage utilisés dans les exports pour une valeur vide.
# À l'import, ils doivent redevenir une chaîne vide.
LIBELLES_VIDES = {"Non traité", "Non traitée", "Non vérifié", "Non vérifiée", "—", "-"}

st.set_page_config(page_title="Cockpit appels — Hympyr", page_icon="📞", layout="wide")

st.markdown(
    f"""
<style>
  h1, h2, h3 {{ color: {VERT_FONCE}; }}
  .stButton>button {{ border-radius: 8px; font-weight: 600; }}
  div[data-testid="stMetricValue"] {{ color: {VERT}; }}
  .fiche {{ background:#f6faf7; border:1px solid #d1e8da; border-left:5px solid {VERT};
           border-radius:10px; padding:16px 20px; margin-bottom:12px; }}
  .pill {{ display:inline-block; background:{VERT}; color:#fff; border-radius:50px;
           padding:3px 13px; font-size:12px; font-weight:600; margin-right:6px; }}
  .pill-orange {{ background:{ORANGE}; }}
  .pill-gris {{ background:{GRIS}; }}
  .bandeau-verrou {{ background:#fff7ed; border:1px solid #f0c48a; border-left:5px solid {ORANGE};
                     border-radius:10px; padding:12px 16px; margin-bottom:12px;
                     font-size:0.92rem; color:#7c3f0a; }}
  .cal-case {{ border:1px solid #d1e8da; border-radius:8px; padding:6px 4px 5px;
               text-align:center; min-height:58px; background:#fff; }}
  .cal-case .j {{ font-size:0.66rem; color:#7a8c85; text-transform:uppercase; letter-spacing:.4px; }}
  .cal-case .d {{ font-size:0.95rem; font-weight:700; color:{VERT_FONCE}; line-height:1.2; }}
  .cal-case .n {{ font-size:0.72rem; font-weight:700; margin-top:2px; }}
  .cal-case.vide {{ background:#fafcfb; border-style:dashed; }}
  .cal-case.hors {{ opacity:.35; }}
  .cal-case.leger  {{ background:#eef7f2; border-color:#bcdfcc; }}
  .cal-case.moyen  {{ background:#d6ecdf; border-color:#8ecfab; }}
  .cal-case.charge {{ background:#{ORANGE.lstrip('#')}22; border-color:{ORANGE}; }}
  .cal-case.retard {{ background:#fdecea; border-color:#e8a49a; }}
  .cal-case.retard .d {{ color:#b4472b; }}
  .cal-case.aujourdhui {{ box-shadow:0 0 0 2px {VERT}; }}
  .cal-pastille {{ display:inline-block; width:8px; height:8px; border-radius:50%; margin:0 1px; }}
  .demo-etape {{ background:{VERT_FONCE}; color:#fff; border-radius:10px;
                 padding:18px 20px; margin-bottom:14px; line-height:1.6; }}
  .demo-etape b {{ color:#8ED6AE; }}
</style>
""",
    unsafe_allow_html=True,
)


# ─────────────────────────────────────────────────────────────────────────────
# UTILITAIRES
# ─────────────────────────────────────────────────────────────────────────────

def empreinte(mot_de_passe: str) -> str:
    """Empreinte SHA-256 salée d'un mot de passe."""
    return hashlib.sha256((SEL + mot_de_passe).encode("utf-8")).hexdigest()


def charger_empreinte(nom: str) -> str:
    """Empreinte attendue pour un profil.

    st.secrets prime sur la valeur inscrite dans le code : cela permet de
    changer un mot de passe sans modifier ni redéployer le fichier source.
    Format attendu dans secrets.toml :  [motsdepasse]  Chloé = "empreinte…"
    """
    try:
        return st.secrets["motsdepasse"][nom]
    except Exception:
        return PROFILS[nom]["empreinte"]


def jours_ouvres(debut: dt.date, fin: dt.date) -> int:
    if fin <= debut:
        return 0
    n, d = 0, debut
    while d < fin:
        if d.weekday() < 5:
            n += 1
        d += dt.timedelta(days=1)
    return n


def date_apres_jours_ouvres(depart: dt.date, nb: int) -> dt.date:
    d, restants = depart, max(0, int(nb))
    while restants > 0:
        d += dt.timedelta(days=1)
        if d.weekday() < 5:
            restants -= 1
    return d


def maintenant_iso() -> str:
    return dt.datetime.now().isoformat(timespec="seconds")


def formater_entier(n) -> str:
    return f"{int(n):,}".replace(",", " ")


def jolie_date(valeur, avec_heure: bool = False) -> str:
    texte = str(valeur or "").strip()
    if not texte:
        return ""
    try:
        d = pd.to_datetime(texte)
        return d.strftime("%d/%m/%Y %H:%M") if avec_heure else d.strftime("%d/%m/%Y")
    except Exception:
        return texte


def vers_iso(valeur) -> str:
    """Convertit une date (JJ/MM/AAAA ou ISO) en ISO, ou chaîne vide si illisible.

    Le format ISO est testé en premier : sans cela, « 2026-09-03 » serait lu en
    jour-mois-année et deviendrait le 9 mars.
    """
    texte = str(valeur or "").strip()
    if not texte:
        return ""
    if re.match(r"^\d{4}-\d{2}-\d{2}", texte):
        try:
            return pd.to_datetime(texte[:10]).date().isoformat()
        except Exception:
            return ""
    for dayfirst in (True, False):
        try:
            return pd.to_datetime(texte, dayfirst=dayfirst).date().isoformat()
        except Exception:
            continue
    return ""


def trouver_colonne(colonnes: Iterable, *cibles: str):
    norm = {str(c).strip().lower(): c for c in colonnes}
    for cible in cibles:
        if cible in norm:
            return norm[cible]
    return None


def pastille_traitement(traite_par: str) -> str:
    """Pastille colorée indiquant qui a traité la fiche, grise si non traitée."""
    nom = str(traite_par or "").strip()
    if nom in PROFILS:
        couleur = PROFILS[nom]["couleur"]
        return (f"<span class='pill' style='background:{couleur}'>Traité par {nom}</span>")
    if nom:
        return f"<span class='pill pill-gris'>Traité par {nom}</span>"
    return "<span class='pill pill-gris'>Non traité</span>"


# ─────────────────────────────────────────────────────────────────────────────
# BASE DE SUIVI (SQLite)
# ─────────────────────────────────────────────────────────────────────────────

@contextmanager
def connexion():
    """Connexion à la base de suivi.

    Le mode WAL n'est PAS réglé ici : c'est un réglage persistant du fichier,
    posé une fois à l'initialisation. L'exécuter à chaque connexion réclamait un
    verrou exclusif momentané — à plusieurs utilisateurs, les connexions
    s'attendaient les unes les autres et l'application semblait figée.
    """
    con = sqlite3.connect(DB_PATH, timeout=30, check_same_thread=False)
    try:
        con.execute("PRAGMA busy_timeout=30000")
        yield con
        con.commit()
    finally:
        con.close()


def initialiser_base() -> None:
    # Réglages persistants du fichier de base : posés une seule fois.
    con = sqlite3.connect(DB_PATH, timeout=30, check_same_thread=False)
    try:
        con.execute("PRAGMA journal_mode=WAL")
        con.execute("PRAGMA synchronous=NORMAL")
        con.commit()
    finally:
        con.close()

    with connexion() as con:
        con.execute("""
            CREATE TABLE IF NOT EXISTS suivi (
                code_client   TEXT PRIMARY KEY,
                statut        TEXT, existe TEXT, produits TEXT,
                email_maj     TEXT, tel_maj TEXT, note TEXT,
                doublon_de    TEXT, rappel_date TEXT, motif_sortie TEXT,
                traite_par    TEXT, maj_le TEXT
            )
        """)
        con.execute("""
            CREATE TABLE IF NOT EXISTS suivi_adresses (
                code_adresse  TEXT PRIMARY KEY,
                referent      TEXT, tel_site TEXT, statut_adr TEXT,
                note_adr      TEXT, traite_par TEXT, maj_le TEXT
            )
        """)
        con.execute("CREATE TABLE IF NOT EXISTS meta (cle TEXT PRIMARY KEY, valeur TEXT)")
        con.execute("""
            CREATE TABLE IF NOT EXISTS journal (
                id          INTEGER PRIMARY KEY AUTOINCREMENT,
                horodatage  TEXT,
                auteur      TEXT,
                action      TEXT,
                cible       TEXT,
                detail      TEXT
            )
        """)
        con.execute("""
            CREATE TABLE IF NOT EXISTS verrous (
                cle        TEXT PRIMARY KEY,
                type_fiche TEXT,
                code       TEXT,
                utilisateur TEXT,
                depuis     TEXT
            )
        """)
        # Migrations sur base existante
        for table, colonnes in (
            ("suivi", ("motif_sortie", "rappel_date", "traite_par")),
            ("suivi_adresses", ("traite_par",)),
        ):
            existantes = {r[1] for r in con.execute(f"PRAGMA table_info({table})").fetchall()}
            for c in colonnes:
                if c not in existantes:
                    con.execute(f"ALTER TABLE {table} ADD COLUMN {c} TEXT")
        con.execute("CREATE INDEX IF NOT EXISTS idx_suivi_maj ON suivi(maj_le)")
        con.execute("CREATE INDEX IF NOT EXISTS idx_suivi_qui ON suivi(traite_par)")
        con.execute("CREATE INDEX IF NOT EXISTS idx_adr_maj ON suivi_adresses(maj_le)")


initialiser_base()


def lire_meta(cle: str, defaut: str = "") -> str:
    with connexion() as con:
        r = con.execute("SELECT valeur FROM meta WHERE cle=?", (cle,)).fetchone()
    return r[0] if r else defaut


def ecrire_meta(cle: str, valeur: str) -> None:
    with connexion() as con:
        con.execute(
            "INSERT INTO meta (cle, valeur) VALUES (?, ?) "
            "ON CONFLICT(cle) DO UPDATE SET valeur=excluded.valeur",
            (cle, str(valeur)),
        )


def charger_suivi() -> pd.DataFrame:
    with connexion() as con:
        return pd.read_sql("SELECT * FROM suivi", con, dtype=str).fillna("")


def charger_suivi_adresses() -> pd.DataFrame:
    with connexion() as con:
        return pd.read_sql("SELECT * FROM suivi_adresses", con, dtype=str).fillna("")


# Champs pour lesquels une valeur vide transmise à l'enregistrement est presque
# toujours accidentelle : une date qui n'a pas pu être relue, un champ non
# réaffiché, un formulaire partiellement rempli. Ils ne sont jamais écrasés par
# du vide, sauf demande explicite d'effacement.
CHAMPS_PROTEGES = ("existe", "produits", "email_maj", "tel_maj", "note",
                   "doublon_de", "rappel_date", "motif_sortie")
CHAMPS_PROTEGES_ADR = ("referent", "tel_site", "note_adr")


def lire_fiche(code: str) -> dict:
    """État actuel d'une fiche en base, ou dictionnaire vide si elle n'existe pas."""
    with connexion() as con:
        colonnes = [r[1] for r in con.execute("PRAGMA table_info(suivi)").fetchall()]
        ligne = con.execute("SELECT * FROM suivi WHERE code_client=?", (str(code),)).fetchone()
    return {c: (v if v is not None else "") for c, v in zip(colonnes, ligne)} if ligne else {}


def lire_fiche_adresse(code: str) -> dict:
    with connexion() as con:
        colonnes = [r[1] for r in con.execute("PRAGMA table_info(suivi_adresses)").fetchall()]
        ligne = con.execute("SELECT * FROM suivi_adresses WHERE code_adresse=?",
                            (str(code),)).fetchone()
    return {c: (v if v is not None else "") for c, v in zip(colonnes, ligne)} if ligne else {}


def _fusionner(existant: dict, champs: dict, proteges: tuple, effacements: Iterable[str]) -> dict:
    """Empêche qu'une valeur déjà saisie soit remplacée par du vide.

    Une donnée inscrite dans une fiche ne disparaît que si son effacement est
    demandé explicitement. C'est ce qui garantit qu'une date de rappel, une note
    ou un e-mail corrigé survivent à tous les enregistrements suivants.
    """
    effacements = set(effacements or ())
    fusion = {}
    for cle, valeur in champs.items():
        texte = "" if valeur is None else str(valeur)
        if (not texte.strip()) and cle in proteges and cle not in effacements and existant.get(cle):
            fusion[cle] = existant[cle]      # on conserve ce qui était déjà saisi
        else:
            fusion[cle] = texte
    return fusion


def enregistrer(code: str, utilisateur: str, effacements: Iterable[str] = (), **champs) -> None:
    """Enregistre une fiche sans jamais perdre une information déjà saisie.

    Les champs listés dans `effacements` sont les seuls que l'on accepte de
    vider. Passer effacements=CHAMPS_PROTEGES rend l'écriture intégrale, ce que
    fait la restauration d'une sauvegarde : elle doit faire foi.
    """
    champs = _fusionner(lire_fiche(code), champs, CHAMPS_PROTEGES, effacements)
    champs["code_client"] = str(code)
    if utilisateur or "traite_par" in effacements:
        champs["traite_par"] = utilisateur
    champs["maj_le"] = maintenant_iso()
    cols = ",".join(champs)
    ph = ",".join("?" for _ in champs)
    upd = ",".join(f"{k}=excluded.{k}" for k in champs if k != "code_client")
    with connexion() as con:
        con.execute(
            f"INSERT INTO suivi ({cols}) VALUES ({ph}) "
            f"ON CONFLICT(code_client) DO UPDATE SET {upd}",
            list(champs.values()),
        )


def enregistrer_adresse(code_adresse: str, utilisateur: str,
                        effacements: Iterable[str] = (), **champs) -> None:
    """Même protection que pour les fiches clients."""
    champs = _fusionner(lire_fiche_adresse(code_adresse), champs,
                        CHAMPS_PROTEGES_ADR, effacements)
    champs["code_adresse"] = str(code_adresse)
    if utilisateur or "traite_par" in effacements:
        champs["traite_par"] = utilisateur
    champs["maj_le"] = maintenant_iso()
    cols = ",".join(champs)
    ph = ",".join("?" for _ in champs)
    upd = ",".join(f"{k}=excluded.{k}" for k in champs if k != "code_adresse")
    with connexion() as con:
        con.execute(
            f"INSERT INTO suivi_adresses ({cols}) VALUES ({ph}) "
            f"ON CONFLICT(code_adresse) DO UPDATE SET {upd}",
            list(champs.values()),
        )


def nb_modifs_depuis_export() -> int:
    ref = lire_meta("dernier_export", "1970-01-01T00:00:00")
    with connexion() as con:
        n = con.execute("SELECT COUNT(*) FROM suivi WHERE maj_le > ?", (ref,)).fetchone()[0]
        n += con.execute("SELECT COUNT(*) FROM suivi_adresses WHERE maj_le > ?", (ref,)).fetchone()[0]
    return int(n)


def base_est_vide() -> bool:
    with connexion() as con:
        n = con.execute("SELECT COUNT(*) FROM suivi").fetchone()[0]
    return n == 0


def reinitialiser_tout(auteur: str = "") -> None:
    """Efface le suivi. Le journal est vidé lui aussi, mais garde une trace de
    l'effacement : sans cela, le bouton annoncerait tout effacer et laisserait
    un historique orphelin renvoyant à des fiches disparues."""
    with connexion() as con:
        for t in ("suivi", "suivi_adresses", "meta", "verrous", "journal"):
            con.execute(f"DELETE FROM {t}")
    if auteur:
        journaliser(auteur, "Réinitialisation complète", "—",
                    "Suivi, référents et journal antérieur effacés.")


def journaliser(auteur: str, action: str, cible: str, detail: str = "") -> None:
    """Trace une intervention manuelle. Sert de preuve en cas de contestation."""
    with connexion() as con:
        con.execute(
            "INSERT INTO journal (horodatage, auteur, action, cible, detail) VALUES (?,?,?,?,?)",
            (maintenant_iso(), auteur, action, str(cible), detail),
        )


def charger_journal(limite: int = 200) -> pd.DataFrame:
    with connexion() as con:
        return pd.read_sql(
            "SELECT horodatage, auteur, action, cible, detail FROM journal "
            "ORDER BY id DESC LIMIT ?", con, params=(limite,), dtype=str
        ).fillna("")


def reattribuer(code: str, traitant: str, statut: str, auteur: str, ancien: dict) -> None:
    """Réassigne manuellement une fiche : qui l'a traitée, et son statut.

    Réservé au profil administrateur. Contrairement à enregistrer(), cette
    fonction n'écrase pas traite_par avec le nom de la personne connectée :
    c'est précisément son objet que de fixer une autre valeur. L'intervention
    est inscrite au journal, avec l'état antérieur.
    """
    with connexion() as con:
        con.execute(
            "INSERT INTO suivi (code_client, statut, traite_par, maj_le) VALUES (?,?,?,?) "
            "ON CONFLICT(code_client) DO UPDATE SET "
            "statut=excluded.statut, traite_par=excluded.traite_par, maj_le=excluded.maj_le",
            (str(code), statut, traitant, maintenant_iso()),
        )
    journaliser(
        auteur, "Réattribution de fiche", code,
        f"traité par « {ancien.get('traite_par') or 'personne'} » → « {traitant or 'personne'} » ; "
        f"statut « {ancien.get('statut') or '—'} » → « {statut} »",
    )


# ── Verrous de consultation ──────────────────────────────────────────────────

def poser_verrou(type_fiche: str, code: str, utilisateur: str) -> None:
    """Signale que cette personne regarde cette fiche, maintenant."""
    with connexion() as con:
        con.execute(
            "INSERT INTO verrous (cle, type_fiche, code, utilisateur, depuis) VALUES (?,?,?,?,?) "
            "ON CONFLICT(cle) DO UPDATE SET utilisateur=excluded.utilisateur, depuis=excluded.depuis",
            (f"{type_fiche}:{code}:{utilisateur}", type_fiche, str(code), utilisateur, maintenant_iso()),
        )
        # Purge occasionnelle seulement : la faire à chaque pose ajoutait un
        # balayage de table à chaque interaction, pour un gain nul.
        if random.random() < 0.05:
            limite = (dt.datetime.now() - dt.timedelta(minutes=VERROU_MINUTES * 3)).isoformat(timespec="seconds")
            con.execute("DELETE FROM verrous WHERE depuis < ?", (limite,))


def marquer_presence(type_fiche: str, code: str, utilisateur: str) -> None:
    """Pose le verrou et mémorise la position, au plus une fois par minute.

    Sans cette limite, chaque clic déclenchait deux écritures en base : le
    verrou et la dernière fiche consultée. Multiplié par trois utilisateurs et
    par le nombre de rafraîchissements de Streamlit, c'était la principale
    source de lenteur.
    """
    cle = type_fiche + ":" + str(code)
    precedent = st.session_state.get("_presence")
    if precedent and precedent[0] == cle and (time.time() - precedent[1]) < 60:
        return
    st.session_state["_presence"] = (cle, time.time())
    poser_verrou(type_fiche, code, utilisateur)
    if type_fiche == "client":
        ecrire_meta("derniere_fiche_" + utilisateur, str(code))


def autres_sur_la_fiche(type_fiche: str, code: str, utilisateur: str) -> list[tuple[str, int]]:
    """Qui d'autre regarde cette fiche, et depuis combien de minutes."""
    limite = (dt.datetime.now() - dt.timedelta(minutes=VERROU_MINUTES)).isoformat(timespec="seconds")
    with connexion() as con:
        lignes = con.execute(
            "SELECT utilisateur, depuis FROM verrous "
            "WHERE type_fiche=? AND code=? AND utilisateur<>? AND depuis >= ?",
            (type_fiche, str(code), utilisateur, limite),
        ).fetchall()
    resultat = []
    for qui, depuis in lignes:
        try:
            minutes = int((dt.datetime.now() - dt.datetime.fromisoformat(depuis)).total_seconds() // 60)
        except Exception:
            minutes = 0
        resultat.append((qui, max(0, minutes)))
    return resultat


def fiches_ouvertes_par_les_autres(type_fiche: str, utilisateur: str) -> set[str]:
    """Codes actuellement consultés par quelqu'un d'autre (verrou actif)."""
    limite = (dt.datetime.now() - dt.timedelta(minutes=VERROU_MINUTES)).isoformat(timespec="seconds")
    with connexion() as con:
        lignes = con.execute(
            "SELECT code FROM verrous WHERE type_fiche=? AND utilisateur<>? AND depuis >= ?",
            (type_fiche, utilisateur, limite),
        ).fetchall()
    return {str(r[0]) for r in lignes}


# ─────────────────────────────────────────────────────────────────────────────
# SAUVEGARDE AUTOMATIQUE
# ─────────────────────────────────────────────────────────────────────────────

def sauvegarde_auto() -> None:
    """Copie CSV du suivi à chaque enregistrement. Ne remplace pas l'export du soir."""
    try:
        jour = dt.date.today().isoformat()
        charger_suivi().to_csv(DOSSIER_SAUVEGARDES / f"suivi_{jour}.csv",
                               index=False, sep=";", encoding="utf-8-sig")
        charger_suivi_adresses().to_csv(DOSSIER_SAUVEGARDES / f"referents_{jour}.csv",
                                        index=False, sep=";", encoding="utf-8-sig")
    except Exception:
        pass  # une sauvegarde qui échoue ne doit jamais bloquer la saisie


# ─────────────────────────────────────────────────────────────────────────────
# IMPORT DES SAUVEGARDES
# ─────────────────────────────────────────────────────────────────────────────

def lire_csv_robuste(fichier) -> pd.DataFrame:
    brut = fichier.read()
    texte = brut.decode("utf-8-sig", errors="replace") if isinstance(brut, bytes) else brut.lstrip("\ufeff")
    premiere = texte.splitlines()[0] if texte.strip() else ""
    sep = ";" if premiere.count(";") >= premiere.count(",") else ","
    return pd.read_csv(io.StringIO(texte), sep=sep, dtype=str).fillna("")


def importer_suivi_clients_csv(fichier) -> int:
    df = lire_csv_robuste(fichier)
    m = {
        "code_client":  trouver_colonne(df.columns, "code_client", "code client"),
        "statut":       trouver_colonne(df.columns, "statut", "statut de l'appel"),
        "existe":       trouver_colonne(df.columns, "existe", "client actif ?"),
        "produits":     trouver_colonne(df.columns, "produits", "produits achetés"),
        "email_maj":    trouver_colonne(df.columns, "email_maj", "e-mail confirmé"),
        "tel_maj":      trouver_colonne(df.columns, "tel_maj", "téléphone confirmé"),
        "doublon_de":   trouver_colonne(df.columns, "doublon_de", "doublon du n°"),
        "motif_sortie": trouver_colonne(df.columns, "motif_sortie", "motif de sortie"),
        "rappel_date":  trouver_colonne(df.columns, "rappel_date", "à rappeler le"),
        "note":         trouver_colonne(df.columns, "note", "notes"),
        "traite_par":   trouver_colonne(df.columns, "traite_par", "traité par", "traitée par"),
    }
    if not m["code_client"]:
        raise ValueError("Le CSV de suivi clients doit contenir une colonne « code_client ».")

    def val(ligne, cle: str) -> str:
        col = m.get(cle)
        brut = str(ligne[col]).strip() if col else ""
        # « Non traité » / « Non vérifié » sont des libellés d'affichage de
        # l'export : ils ne doivent pas revenir en base comme un nom de personne.
        return "" if brut in LIBELLES_VIDES else brut

    n = 0
    for _, ligne in df.iterrows():
        code = val(ligne, "code_client")
        if not code:
            continue
        produits = val(ligne, "produits").replace(";", ",")
        produits = "|".join(x.strip() for x in produits.split(",") if x.strip())
        enregistrer(
            code,
            utilisateur=val(ligne, "traite_par"),   # on conserve l'auteur d'origine
            # La restauration fait foi : elle réécrit la fiche telle qu'exportée,
            # y compris les champs vides.
            effacements=CHAMPS_PROTEGES + ("traite_par",),
            statut=val(ligne, "statut") or "À appeler",
            existe=val(ligne, "existe"),
            produits=produits,
            email_maj=val(ligne, "email_maj"),
            tel_maj=val(ligne, "tel_maj"),
            doublon_de=val(ligne, "doublon_de"),
            motif_sortie=val(ligne, "motif_sortie"),
            note=val(ligne, "note"),
            rappel_date=vers_iso(val(ligne, "rappel_date")),
        )
        n += 1
    return n


def importer_suivi_adresses_csv(fichier) -> int:
    df = lire_csv_robuste(fichier)
    m = {
        "code_adresse": trouver_colonne(df.columns, "code_adresse", "code adresse"),
        "referent":     trouver_colonne(df.columns, "referent", "référent sur place"),
        "tel_site":     trouver_colonne(df.columns, "tel_site", "tél. référent / site", "tel. référent / site"),
        "statut_adr":   trouver_colonne(df.columns, "statut_adr", "statut vérification"),
        "note_adr":     trouver_colonne(df.columns, "note_adr", "note"),
        "traite_par":   trouver_colonne(df.columns, "traite_par", "vérifié par", "traité par"),
    }
    if not m["code_adresse"]:
        raise ValueError("Le CSV des référents doit contenir une colonne « code_adresse ».")

    def val(ligne, cle: str) -> str:
        col = m.get(cle)
        brut = str(ligne[col]).strip() if col else ""
        return "" if brut in LIBELLES_VIDES else brut

    n = 0
    for _, ligne in df.iterrows():
        code = val(ligne, "code_adresse")
        if not code:
            continue
        # Une ligne sans aucune saisie ne crée pas d'enregistrement : l'export
        # des points de livraison contient tous les sites du fichier mère,
        # y compris ceux que personne n'a encore vérifiés.
        if not any(val(ligne, c) for c in ("referent", "tel_site", "note_adr", "traite_par")) \
                and val(ligne, "statut_adr") in ("", "À vérifier"):
            continue
        enregistrer_adresse(
            code,
            utilisateur=val(ligne, "traite_par"),
            effacements=CHAMPS_PROTEGES_ADR + ("traite_par",),
            referent=val(ligne, "referent"),
            tel_site=val(ligne, "tel_site"),
            statut_adr=val(ligne, "statut_adr") or "À vérifier",
            note_adr=val(ligne, "note_adr"),
        )
        n += 1
    return n


# ─────────────────────────────────────────────────────────────────────────────
# FICHIER MÈRE — conservé au niveau du processus, pas de la session
# ─────────────────────────────────────────────────────────────────────────────

@st.cache_resource(show_spinner=False)
def coffre_fichier() -> dict:
    """Conteneur partagé par toutes les sessions du processus.

    st.session_state est détruit dès que le websocket tombe (mise en veille,
    onglet en arrière-plan, changement de réseau) ; un cache de ressource
    survit à la reconnexion. C'est ce qui évite que l'outil « se réinitialise
    tout seul ».
    """
    return {"contenu": None, "nom": "", "charge_le": "", "restaure": False,
            "clients": None, "adresses": None, "col_code": None}


def lire_fichier(contenu: bytes):
    tampon = io.BytesIO(contenu)
    xls = pd.ExcelFile(tampon, engine="openpyxl")
    clients = pd.read_excel(xls, "Clients", dtype=str).fillna("")
    adresses = (
        pd.read_excel(xls, "Adresses livraison", dtype=str).fillna("")
        if "Adresses livraison" in xls.sheet_names else pd.DataFrame()
    )
    return clients, adresses


def normaliser_clients(clients: pd.DataFrame):
    col_code = trouver_colonne(clients.columns, "code client", "code_client")
    if col_code is None:
        return clients, None

    renoms = {}
    for sources, cible in [
        (("raison sociale / nom", "nom", "raison sociale"), "Raison sociale / Nom"),
        (("type client", "type"), "Type client"),
        (("ville",), "Ville"),
        (("catégorie normalisée", "catégorie", "categorie"), "Catégorie"),
        (("siren (9 chiffres)", "siren", "siren / siret"), "SIREN"),
        (("code postal", "code_postal"), "Code postal"),
    ]:
        col = trouver_colonne(clients.columns, *sources)
        if col and col != cible:
            renoms[col] = cible
    for n in (1, 2, 3):
        col = trouver_colonne(clients.columns, f"téléphone {n} (norm)", f"téléphone {n}", f"telephone {n}")
        if col and col != f"Téléphone {n}":
            renoms[col] = f"Téléphone {n}"
    if renoms:
        clients = clients.rename(columns=renoms)
    clients = clients.loc[:, ~clients.columns.duplicated(keep="first")]

    for c in ["Raison sociale / Nom", "Type client", "Ville", "Catégorie", "À compléter",
              "Email principal", "Email secondaire", "SIREN",
              "Téléphone 1", "Téléphone 2", "Téléphone 3",
              "Adresse 1", "Adresse 2", "Adresse 3", "Code postal"]:
        if c not in clients.columns:
            clients[c] = ""
    return clients, col_code


def normaliser_adresses(adresses: pd.DataFrame) -> pd.DataFrame:
    if adresses.empty:
        return adresses
    col = trouver_colonne(adresses.columns, "code client mère", "code client mere")
    if col and col != "Code client mère":
        adresses = adresses.rename(columns={col: "Code client mère"})
    if "Code adresse" not in adresses.columns:
        col = trouver_colonne(adresses.columns, "code adresse") or trouver_colonne(adresses.columns, "code client")
        if col:
            adresses = adresses.rename(columns={col: "Code adresse"})
    for sources, cible in [
        (("téléphone 1 (norm)", "téléphone 1", "telephone 1"), "Téléphone"),
        (("code postal",), "Code postal"),
        (("nom site", "nom"), "Nom site"),
    ]:
        col = trouver_colonne(adresses.columns, *sources)
        if col and col != cible and cible not in adresses.columns:
            adresses = adresses.rename(columns={col: cible})
    adresses = adresses.loc[:, ~adresses.columns.duplicated(keep="first")]
    for c in ["Code adresse", "Code client mère", "Nom site", "Adresse 1", "Adresse 2",
              "Adresse 3", "Code postal", "Ville", "Téléphone"]:
        if c not in adresses.columns:
            adresses[c] = ""
    return adresses


def preparer_fichier() -> None:
    """Lit et normalise le fichier une seule fois, à son chargement.

    Auparavant, chaque rafraîchissement repassait par le cache de Streamlit en
    lui donnant le contenu du fichier comme clé : plusieurs mégaoctets à
    empreinter à chaque interaction, puis une renormalisation complète des
    colonnes. Le résultat est désormais conservé avec le fichier lui-même.
    """
    clients_bruts, adresses_brutes = lire_fichier(coffre["contenu"])
    clients, col_code = normaliser_clients(clients_bruts)
    coffre["clients"] = clients
    coffre["col_code"] = col_code
    coffre["adresses"] = normaliser_adresses(adresses_brutes)


def priorite(type_client: str) -> int:
    """Ordre d'appel : pros d'abord (enjeu de conformité), particuliers en dernier."""
    t = (type_client or "").lower()
    if t.startswith("pro") and "déduit" not in t:
        return 0
    if t.startswith("pro"):
        return 1
    if "déterminer" in t:
        return 2
    if "public" in t or "asso" in t:
        return 3
    return 4


# ─────────────────────────────────────────────────────────────────────────────
# EXPORTS
# ─────────────────────────────────────────────────────────────────────────────

def preparer_export(suivi_df: pd.DataFrame, base_df: pd.DataFrame, col_code: str) -> pd.DataFrame:
    df = suivi_df.copy().fillna("")
    if df.empty:
        return df

    infos = base_df[[col_code, "Raison sociale / Nom", "Ville", "Type client"]].copy()
    infos = infos.rename(columns={col_code: "code_client"})
    df = df.merge(infos, on="code_client", how="left").fillna("")

    df["produits"] = df["produits"].str.replace("|", ", ", regex=False)
    df["rappel_date"] = df["rappel_date"].map(lambda v: jolie_date(v))
    df["maj_le"] = df["maj_le"].map(lambda v: jolie_date(v, avec_heure=True))
    df["traite_par"] = df["traite_par"].replace("", "Non traité")

    colonnes = {
        "code_client": "Code client",
        "Raison sociale / Nom": "Nom / Raison sociale",
        "Ville": "Ville",
        "Type client": "Type",
        "traite_par": "Traité par",
        "statut": "Statut de l'appel",
        "existe": "Client actif ?",
        "produits": "Produits achetés",
        "email_maj": "E-mail confirmé",
        "tel_maj": "Téléphone confirmé",
        "doublon_de": "Doublon du n°",
        "motif_sortie": "Motif de sortie",
        "rappel_date": "À rappeler le",
        "note": "Notes",
        "maj_le": "Dernière mise à jour",
    }
    for c in colonnes:
        if c not in df.columns:
            df[c] = ""
    return df[list(colonnes)].rename(columns=colonnes)


def preparer_export_adresses(suivi_adr: pd.DataFrame, adresses_df: pd.DataFrame) -> pd.DataFrame:
    if adresses_df.empty:
        return pd.DataFrame()

    sortie = adresses_df[["Code adresse", "Code client mère", "Nom site", "Code postal", "Ville"]].rename(
        columns={"Nom site": "Nom du site"}
    )
    if suivi_adr.empty:
        suivi_adr = pd.DataFrame(columns=["code_adresse", "referent", "tel_site",
                                          "statut_adr", "note_adr", "traite_par", "maj_le"])
    sortie = sortie.merge(suivi_adr, left_on="Code adresse", right_on="code_adresse", how="left").fillna("")
    sortie["maj_le"] = sortie.get("maj_le", "").map(lambda v: jolie_date(v, avec_heure=True))
    sortie["statut_adr"] = sortie.get("statut_adr", "").replace("", "À vérifier")
    sortie["traite_par"] = sortie.get("traite_par", "").replace("", "Non traité")

    libelles = {
        "traite_par": "Vérifié par",
        "referent": "Référent sur place",
        "tel_site": "Tél. référent / site",
        "statut_adr": "Statut vérification",
        "note_adr": "Note",
        "maj_le": "Dernière mise à jour",
    }
    for c in libelles:
        if c not in sortie.columns:
            sortie[c] = ""
    ordre = ["Code adresse", "Code client mère", "Nom du site", "Code postal", "Ville",
             "traite_par", "referent", "tel_site", "statut_adr", "note_adr", "maj_le"]
    return sortie[ordre].rename(columns=libelles)


def vers_excel(df: pd.DataFrame, nom_feuille: str) -> bytes:
    from openpyxl.styles import Alignment, Font, PatternFill

    tampon = io.BytesIO()
    with pd.ExcelWriter(tampon, engine="openpyxl") as writer:
        df.to_excel(writer, index=False, sheet_name=nom_feuille)
        ws = writer.sheets[nom_feuille]
        for j, col in enumerate(df.columns, 1):
            cellule = ws.cell(row=1, column=j)
            cellule.fill = PatternFill("solid", fgColor=VERT_FONCE.lstrip("#"))
            cellule.font = Font(bold=True, color="FFFFFF")
            cellule.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            longueurs = df[col].astype(str).str.len().head(200)
            ws.column_dimensions[cellule.column_letter].width = min(
                max(len(str(col)) + 2, int(longueurs.max() or 10) + 2), 45)
        ws.freeze_panes = "A2"
        ws.auto_filter.ref = ws.dimensions
    return tampon.getvalue()


# ─────────────────────────────────────────────────────────────────────────────
# INDICATEURS
# ─────────────────────────────────────────────────────────────────────────────

def calculer_rythme(suivi_df: pd.DataFrame, utilisateur: str | None = None) -> float | None:
    """Fiches terminées par jour d'activité réelle, globalement ou par personne."""
    if suivi_df.empty or "maj_le" not in suivi_df.columns:
        return None
    termines = suivi_df[suivi_df["statut"].isin(STATUTS_TERMINES)].copy()
    if utilisateur:
        termines = termines[termines["traite_par"] == utilisateur]
    if termines.empty:
        return None
    termines["jour"] = pd.to_datetime(termines["maj_le"], errors="coerce").dt.date
    jours_actifs = termines["jour"].nunique()
    return len(termines) / jours_actifs if jours_actifs else None


def rappels_dus(base_df: pd.DataFrame, utilisateur: str | None = None) -> pd.DataFrame:
    """Clients à rappeler aujourd'hui ou dont la date de rappel est dépassée."""
    if base_df.empty:
        return base_df
    df = base_df[(base_df["statut"] == "À rappeler")
                 & (base_df["rappel_date"].astype(str).str.strip() != "")].copy()
    if df.empty:
        return df
    df["date_rappel"] = pd.to_datetime(df["rappel_date"], errors="coerce").dt.date
    df = df[df["date_rappel"].notna() & (df["date_rappel"] <= dt.date.today())]
    if utilisateur:
        df = df[df["traite_par"] == utilisateur]
    return df.sort_values("date_rappel")


def rappels_planifies(base_df: pd.DataFrame, utilisateur: str | None = None) -> pd.DataFrame:
    """Toutes les fiches portant une date de rappel, passées comme à venir.

    Même source que le bandeau d'alerte : le champ rempli par les commerciales
    dans le formulaire de résultat d'appel. Le calendrier ne fait que le lire.
    """
    if base_df.empty:
        return base_df
    df = base_df[(base_df["statut"] == "À rappeler")
                 & (base_df["rappel_date"].astype(str).str.strip() != "")].copy()
    if df.empty:
        return df
    df["date_rappel"] = pd.to_datetime(df["rappel_date"], errors="coerce").dt.date
    df = df[df["date_rappel"].notna()]
    if utilisateur:
        df = df[df["traite_par"] == utilisateur]
    return df.sort_values("date_rappel")


def _classe_charge(nombre: int, jour: dt.date) -> str:
    """Couleur d'une case selon la charge, le retard primant sur le volume."""
    if nombre == 0:
        return "vide"
    if jour < dt.date.today():
        return "retard"
    if nombre >= 8:
        return "charge"
    if nombre >= 4:
        return "moyen"
    return "leger"


def _pastilles(rappels_du_jour: pd.DataFrame) -> str:
    """Une pastille par commerciale concernée, pour lire la répartition d'un coup."""
    if rappels_du_jour.empty:
        return ""
    points = ""
    for nom in COMMERCIALES:
        n = int((rappels_du_jour["traite_par"] == nom).sum())
        if n:
            points += ("<span class='cal-pastille' title='" + nom + " : " + str(n)
                       + "' style='background:" + PROFILS[nom]["couleur"] + "'></span>")
    return points


JOURS_SEMAINE = ["lun", "mar", "mer", "jeu", "ven", "sam", "dim"]
MOIS_LONGS = ["janvier", "février", "mars", "avril", "mai", "juin",
              "juillet", "août", "septembre", "octobre", "novembre", "décembre"]


def bande_sept_jours(rappels: pd.DataFrame, prefixe: str) -> None:
    """Sept prochains jours, au-dessus de la file d'appel.

    Format volontairement compact : les commerciales travaillent à la semaine,
    et cette bande ne doit pas repousser l'outil qu'elles utilisent réellement
    hors de l'écran. Cliquer sur un jour filtre la file sur ses rappels.
    """
    aujourdhui = dt.date.today()
    compte = rappels.groupby("date_rappel").size().to_dict() if not rappels.empty else {}
    retard = sum(n for j, n in compte.items() if j < aujourdhui)

    entete = "**Vos sept prochains jours**"
    if retard:
        entete += "  ·  :red[" + str(retard) + " rappel(s) en retard]"
    st.markdown(entete)

    colonnes = st.columns(7)
    for i, col in enumerate(colonnes):
        jour = aujourdhui + dt.timedelta(days=i)
        n = int(compte.get(jour, 0))
        classes = "cal-case " + _classe_charge(n, jour) + (" aujourdhui" if i == 0 else "")
        col.markdown(
            "<div class='" + classes + "'>"
            "<div class='j'>" + JOURS_SEMAINE[jour.weekday()] + "</div>"
            "<div class='d'>" + str(jour.day) + "</div>"
            "<div class='n'>" + (str(n) if n else "—") + "</div></div>",
            unsafe_allow_html=True,
        )
        if col.button("Voir" if n else "—", key=prefixe + "_" + jour.isoformat(),
                      disabled=(n == 0), use_container_width=True):
            st.session_state.jour_rappel = jour.isoformat()
            st.session_state.pop("idx", None)
            st.rerun()

    if retard:
        if st.button("⏰ Voir les " + str(retard) + " rappel(s) en retard", key=prefixe + "_retard"):
            st.session_state.jour_rappel = "retard"
            st.session_state.pop("idx", None)
            st.rerun()


def grille_mensuelle(rappels: pd.DataFrame, prefixe: str) -> None:
    """Grille d'un mois entier, pour le pilotage.

    Une case par jour : le nombre de rappels, et une pastille par commerciale
    concernée. C'est la charge à venir, que rien d'autre dans l'outil ne montre.
    """
    if "cal_decalage" not in st.session_state:
        st.session_state.cal_decalage = 0

    aujourdhui = dt.date.today()
    mois = aujourdhui.month - 1 + st.session_state.cal_decalage
    annee = aujourdhui.year + mois // 12
    mois = mois % 12 + 1

    n1, n2, n3 = st.columns([1, 3, 1])
    if n1.button("◀ Mois précédent", key=prefixe + "_prec", use_container_width=True):
        st.session_state.cal_decalage -= 1
        st.rerun()
    if n3.button("Mois suivant ▶", key=prefixe + "_suiv", use_container_width=True):
        st.session_state.cal_decalage += 1
        st.rerun()
    n2.markdown(
        "<div style='text-align:center;font-weight:700;color:" + VERT_FONCE +
        ";padding-top:6px'>" + MOIS_LONGS[mois - 1] + " " + str(annee) + "</div>",
        unsafe_allow_html=True,
    )

    par_jour = {}
    if not rappels.empty:
        for jour, groupe in rappels.groupby("date_rappel"):
            par_jour[jour] = groupe

    premier = dt.date(annee, mois, 1)
    dernier = dt.date(annee + (mois == 12), mois % 12 + 1, 1) - dt.timedelta(days=1)
    debut = premier - dt.timedelta(days=premier.weekday())      # lundi de la 1re semaine

    st.markdown(
        "<div style='display:grid;grid-template-columns:repeat(7,1fr);gap:4px;"
        "text-align:center;font-size:0.7rem;color:#7a8c85;text-transform:uppercase;"
        "letter-spacing:.4px;margin-bottom:4px'>"
        + "".join("<div>" + j + "</div>" for j in JOURS_SEMAINE) + "</div>",
        unsafe_allow_html=True,
    )

    cases = []
    jour = debut
    while jour <= dernier or jour.weekday() != 0:
        groupe = par_jour.get(jour)
        n = 0 if groupe is None else len(groupe)
        hors = not (premier <= jour <= dernier)
        classes = "cal-case " + _classe_charge(n, jour)
        if hors:
            classes += " hors"
        if jour == aujourdhui:
            classes += " aujourdhui"
        cases.append(
            "<div class='" + classes + "'>"
            "<div class='d'>" + str(jour.day) + "</div>"
            "<div class='n'>" + (str(n) if n else "&nbsp;") + "</div>"
            "<div>" + (_pastilles(groupe) if groupe is not None else "") + "</div></div>"
        )
        jour += dt.timedelta(days=1)
        if len(cases) > 42:
            break

    st.markdown(
        "<div style='display:grid;grid-template-columns:repeat(7,1fr);gap:4px'>"
        + "".join(cases) + "</div>",
        unsafe_allow_html=True,
    )

    legende = "  ·  ".join(
        "<span class='cal-pastille' style='background:" + PROFILS[nom]["couleur"] + "'></span> " + nom
        for nom in COMMERCIALES
    )
    st.markdown(
        "<div style='font-size:0.76rem;color:#5a6b64;margin-top:10px'>" + legende +
        "  ·  Une case rouge signale des rappels dont la date est dépassée.</div>",
        unsafe_allow_html=True,
    )

    du_mois = rappels[(rappels["date_rappel"] >= premier) & (rappels["date_rappel"] <= dernier)] \
        if not rappels.empty else rappels
    if du_mois.empty:
        st.caption("Aucun rappel programmé sur ce mois.")
        return
    st.caption(str(len(du_mois)) + " rappel(s) programmé(s) sur le mois affiché.")

    # Sous la grille : un bouton par journée concernée, puis la liste des
    # clients. La couleur d'une case montre une charge ; seule cette liste
    # permet de décrocher le téléphone.
    st.divider()
    jour_choisi = selecteur_journees(rappels, premier, dernier, prefixe)
    if jour_choisi:
        du_jour = rappels[rappels["date_rappel"] == jour_choisi]
        titre = str(len(du_jour)) + " client(s) à rappeler le " + jour_choisi.strftime("%d/%m/%Y")
        if jour_choisi < dt.date.today():
            titre += "  —  date dépassée"
        tableau_rappels(du_jour, col_code or "Code client", titre)
        detail = []
        for nom in COMMERCIALES:
            n = int((du_jour["traite_par"] == nom).sum())
            if n:
                detail.append(nom + " : " + str(n))
        if detail:
            st.caption("Répartition — " + "  ·  ".join(detail) + ".")


def tableau_rappels(rappels: pd.DataFrame, col_code: str, titre: str = "") -> None:
    """Liste des clients à rappeler : qui appeler, à quel numéro, et pourquoi.

    C'est ce que le calendrier doit produire. Une case colorée indique une
    charge ; seule cette liste permet de décrocher le téléphone.
    """
    if rappels.empty:
        st.info("Aucun rappel pour cette date.")
        return
    if titre:
        st.markdown("**" + titre + "**")

    vue = rappels.copy()
    vue["Date de rappel"] = vue["date_rappel"].map(lambda d: d.strftime("%d/%m/%Y"))
    vue["Traité par"] = vue["traite_par"].replace("", "—")
    colonnes = {
        col_code: "Code client",
        "Raison sociale / Nom": "Client",
        "Ville": "Ville",
        "Téléphone 1": "Téléphone",
        "Traité par": "Noté par",
        "Date de rappel": "Date de rappel",
        "note": "Notes de l'appel précédent",
    }
    for c in colonnes:
        if c not in vue.columns:
            vue[c] = ""
    st.dataframe(vue[list(colonnes)].rename(columns=colonnes),
                 hide_index=True, use_container_width=True)


def selecteur_journees(rappels: pd.DataFrame, premier: dt.date, dernier: dt.date,
                       prefixe: str) -> dt.date | None:
    """Boutons de sélection, un par journée comportant des rappels.

    Seules les journées concernées ont un bouton : sur un mois, cela fait une
    poignée d'éléments au lieu des quarante-deux cases de la grille.
    """
    if rappels.empty:
        return None
    jours = sorted({d for d in rappels["date_rappel"] if premier <= d <= dernier})
    if not jours:
        return None

    cle_etat = prefixe + "_jour_detail"
    st.markdown("**Journées comportant des rappels — cliquez pour voir les clients**")
    par_ligne = 6
    for debut in range(0, len(jours), par_ligne):
        colonnes = st.columns(par_ligne)
        for col, jour in zip(colonnes, jours[debut:debut + par_ligne]):
            n = int((rappels["date_rappel"] == jour).sum())
            marque = "⏰ " if jour < dt.date.today() else ""
            libelle = marque + jour.strftime("%d/%m") + " · " + str(n)
            actif = st.session_state.get(cle_etat) == jour.isoformat()
            if col.button(libelle, key=prefixe + "_j_" + jour.isoformat(),
                          use_container_width=True,
                          type="primary" if actif else "secondary"):
                # Un second clic sur la même journée referme le détail.
                if actif:
                    st.session_state.pop(cle_etat, None)
                else:
                    st.session_state[cle_etat] = jour.isoformat()
                st.rerun()

    choisi = st.session_state.get(cle_etat)
    if choisi:
        try:
            return dt.date.fromisoformat(choisi)
        except ValueError:
            st.session_state.pop(cle_etat, None)
    return None


def stats_utilisateur(suivi_df: pd.DataFrame, suivi_adr: pd.DataFrame, nom: str) -> dict:
    """Indicateurs d'activité d'une commerciale."""
    aujourdhui = dt.date.today()
    s = suivi_df[suivi_df["traite_par"] == nom] if not suivi_df.empty else pd.DataFrame()
    a = suivi_adr[suivi_adr["traite_par"] == nom] if not suivi_adr.empty else pd.DataFrame()

    def du_jour(df):
        if df.empty:
            return 0
        jours = pd.to_datetime(df["maj_le"], errors="coerce").dt.date
        return int((jours == aujourdhui).sum())

    return {
        "fiches": len(s),
        "terminees": int(s["statut"].isin(STATUTS_TERMINES).sum()) if not s.empty else 0,
        "aujourdhui": du_jour(s),
        "a_rappeler": int((s["statut"] == "À rappeler").sum()) if not s.empty else 0,
        "points": len(a),
        "points_aujourdhui": du_jour(a),
        "rythme": calculer_rythme(suivi_df, nom),
    }


# ─────────────────────────────────────────────────────────────────────────────
# DÉMO GUIDÉE
# ─────────────────────────────────────────────────────────────────────────────

DEMO_COMMERCIAL = [
    ("La colonne de gauche : votre poste de pilotage",
     "Elle affiche <b>vos chiffres du jour</b> et l'avancement de l'équipe, puis les "
     "<b>filtres</b> qui construisent votre file d'appel. Le filtre le plus important est "
     "« Traitement » : choisissez <b>Non traité</b> pour ne voir que les fiches que personne "
     "n'a encore prises. C'est ce qui vous évite d'appeler un client déjà appelé par votre collègue."),
    ("Onglet « Appels clients » : une fiche à la fois",
     "À gauche, les coordonnées du client. À droite, le formulaire de résultat d'appel. "
     "Une <b>pastille colorée</b> en haut indique qui a déjà traité la fiche — grise si personne. "
     "Si votre collègue est en train de la consulter, un <b>bandeau orange</b> vous prévient : "
     "passez simplement à la suivante."),
    ("Onglet « Points de livraison » : le prolongement de l'appel",
     "Une entreprise cliente peut avoir plusieurs sites livrés. Sur la fiche client, un bloc "
     "<b>« adresses de livraison rattachées »</b> vous les montre. "
     "Pendant l'appel, notez le <b>référent sur place</b> pour chaque site, puis reportez-le dans "
     "cet onglet — la recherche par <b>code client mère</b> affiche tous les points d'un même client."),
    ("Le déroulé d'un appel, en cinq gestes",
     "1. Filtre <b>Non traité</b> dans la colonne de gauche.<br>"
     "2. Vous appelez le client affiché.<br>"
     "3. Vous remplissez le formulaire de droite : statut, produits, e-mail et téléphone confirmés.<br>"
     "4. Vous cliquez sur <b>Enregistrer & passer au suivant</b> — la fiche suivante s'ouvre seule.<br>"
     "5. Si le client a plusieurs sites, vous complétez l'onglet <b>Points de livraison</b>."),
    ("Onglet « Export » : à faire avant de partir",
     "Un compteur vous indique le nombre de modifications non exportées. "
     "En fin de journée, téléchargez les <b>deux fichiers</b> — suivi clients et points de livraison — "
     "puis confirmez pour repasser au vert. C'est votre sauvegarde."),
]

DEMO_MANAGER = [
    ("La colonne de gauche : la vue d'ensemble",
     "Vous y trouvez l'avancement global de la campagne, le <b>rythme observé</b>, la projection de "
     "fin, et l'objectif quotidien à tenir pour l'échéance de la facturation électronique. "
     "Juste en dessous, les <b>chiffres de Chloé et de Patricia</b>, mis à jour à chaque "
     "enregistrement de leur part."),
    ("Tableau de bord : tenir l'échéance",
     "Le premier bloc compare l'<b>objectif quotidien</b> au <b>rythme réel</b>. Vous pouvez "
     "restreindre le périmètre aux seuls clients professionnels, qui sont ceux concernés par "
     "l'obligation. Si le rythme passe sous l'objectif, une projection de fin s'affiche en rouge."),
    ("Performance par commerciale",
     "Un tableau et des graphiques comparent l'activité de Chloé et de Patricia : fiches traitées, "
     "fiches du jour, rappels en cours, points de livraison vérifiés. "
     "Ces données viennent <b>directement de leur saisie</b>, sans ressaisie ni déclaratif."),
    ("Ce que vous voyez, et ce que vous ne voyez pas",
     "Votre profil est en <b>lecture seule</b> : vous ne pouvez ni modifier une fiche, ni exporter "
     "à leur place. C'est volontaire — cela garantit que les chiffres reflètent uniquement le "
     "travail réellement effectué."),
]


DEMO_ADMIN = [
    ("Votre profil : accès complet",
     "Vous disposez des <b>quatre onglets</b> : les trois onglets de travail des commerciales "
     "(appels, points de livraison, export) et le <b>tableau de bord</b> du manager. "
     "Seule la <b>réinitialisation</b> vous est fermée : elle reste exclusive au profil manager, "
     "pour qu'un effacement total ne puisse jamais partir d'un poste qui saisit."),
    ("La colonne de gauche : vos chiffres et ceux de l'équipe",
     "Elle cumule les deux vues : <b>vos propres chiffres</b> du jour, l'avancement global, "
     "et les <b>chiffres de Chloé et de Patricia</b>. En dessous, les filtres qui construisent "
     "votre file d'appel — dont « Traitement », qui évite les doublons."),
    ("Onglets de travail : comme les commerciales, avec un pouvoir en plus",
     "Sur une fiche, une <b>pastille colorée</b> indique qui l'a traitée. Vos propres "
     "enregistrements portent votre nom, en bleu. Un <b>bandeau orange</b> vous prévient si "
     "quelqu'un consulte la même fiche en même temps que vous.<br><br>"
     "Sous le formulaire, un bloc <b>« Administration de la fiche »</b> n'apparaît que pour vous : "
     "il permet de <b>réattribuer</b> une fiche à quelqu'un d'autre, de la <b>rendre à la file</b> "
     "en la repassant à « non traité », ou de <b>forcer un statut</b>. "
     "Chaque intervention est inscrite au journal, consultable depuis le tableau de bord."),
    ("Onglet « Tableau de bord » : la vue de pilotage",
     "Objectif quotidien face au rythme réel, performance comparée des deux commerciales, "
     "avancement global et rappels en retard. Vos propres fiches sont comptées dans l'avancement "
     "global, et signalées à part dans le bloc de performance : le suivi des commerciales reste lisible."),
    ("Ce qu'il faut faire chaque soir",
     "Comme les commerciales, pensez à l'onglet <b>Export</b> : un compteur indique les "
     "modifications non sauvegardées. Les deux fichiers téléchargés servent à rouvrir "
     "une session le lendemain."),
]


def afficher_demo(role: str) -> None:
    """Visite guidée en pop-up, proposée à l'ouverture de session."""
    etapes = {"commercial": DEMO_COMMERCIAL, "manager": DEMO_MANAGER, "admin": DEMO_ADMIN}[role]
    idx = st.session_state.get("demo_etape", 0)
    idx = max(0, min(idx, len(etapes) - 1))
    titre, texte = etapes[idx]

    def contenu():
        st.markdown(f"<div class='demo-etape'><b>Étape {idx + 1} / {len(etapes)}</b><br><br>"
                    f"<span style='font-size:1.05rem;font-weight:700'>{titre}</span><br><br>"
                    f"{texte}</div>", unsafe_allow_html=True)
        st.progress((idx + 1) / len(etapes))
        c1, c2, c3 = st.columns([1, 1, 1])
        if c1.button("⬅️ Précédent", disabled=(idx == 0), use_container_width=True, key="demo_prec"):
            st.session_state.demo_etape = idx - 1
            st.rerun()
        if idx < len(etapes) - 1:
            if c2.button("Suivant ➡️", type="primary", use_container_width=True, key="demo_suiv"):
                st.session_state.demo_etape = idx + 1
                st.rerun()
        else:
            if c2.button("Terminer", type="primary", use_container_width=True, key="demo_fin"):
                st.session_state.demo_active = False
                st.rerun()
        if c3.button("Fermer", use_container_width=True, key="demo_close"):
            st.session_state.demo_active = False
            st.rerun()

    # Rendu dans le flux de la page plutôt que dans une fenêtre modale.
    # st.dialog était rappelé à chaque exécution du script tant que la démo
    # restait ouverte, et chacun de ses boutons relançait le script : sur un
    # serveur chargé, l'application tournait en boucle sans jamais se stabiliser.
    with st.container(border=True):
        st.subheader("Visite guidée de l'outil")
        contenu()


# ─────────────────────────────────────────────────────────────────────────────
# ÉCRAN DE CONNEXION
# ─────────────────────────────────────────────────────────────────────────────

coffre = coffre_fichier()


def ecran_connexion() -> None:
    """Chargement des fichiers, puis choix du profil et mot de passe."""
    st.title("Cockpit appels — Hympyr Énergies")

    # ── Étape 1 : les trois fichiers ─────────────────────────────────────────
    if coffre["contenu"] is None or not coffre["restaure"]:
        st.subheader("1. Charger les fichiers de travail")
        st.caption(
            "Les trois fichiers sont nécessaires pour ouvrir une session : le fichier clients "
            "et les deux sauvegardes exportées lors de la dernière session. "
            "Une fois chargés, ils restent disponibles pour toute l'équipe jusqu'au redémarrage du serveur."
        )

        c1, c2, c3 = st.columns(3)
        f_mere = c1.file_uploader("Fichier clients (.xlsx)", type=["xlsx"], key="up_mere")
        f_suivi = c2.file_uploader("Sauvegarde SUIVI CLIENTS (.csv)", type=["csv"], key="up_suivi")
        f_refs = c3.file_uploader("Sauvegarde RÉFÉRENTS (.csv)", type=["csv"], key="up_refs")

        deja_charge = coffre["contenu"] is not None
        if deja_charge:
            st.success(f"Fichier clients **{coffre['nom']}** déjà en mémoire.")

        # Échappatoire si la base contient déjà le travail : demander les deux
        # sauvegardes serait alors un risque de régression, pas une sécurité.
        reprendre = False
        if not base_est_vide():
            with connexion() as con:
                nb = con.execute("SELECT COUNT(*) FROM suivi").fetchone()[0]
            reprendre = st.checkbox(
                f"La base contient déjà {formater_entier(nb)} fiche(s) — reprendre sans restaurer les sauvegardes",
                help="À cocher uniquement si le travail en cours est déjà dans l'outil. "
                     "Restaurer une sauvegarde plus ancienne écraserait les saisies récentes.",
            )

        pret = (f_mere is not None or deja_charge) and (reprendre or (f_suivi is not None and f_refs is not None))
        if st.button("Ouvrir la session", type="primary", disabled=not pret):
            try:
                if f_mere is not None:
                    coffre["contenu"] = f_mere.getvalue()
                    coffre["nom"] = f_mere.name
                    coffre["charge_le"] = maintenant_iso()
                    coffre["clients"] = None      # sera préparé au premier affichage
                messages = []
                if not reprendre:
                    messages.append(f"{importer_suivi_clients_csv(f_suivi)} fiches clients restaurées")
                    messages.append(f"{importer_suivi_adresses_csv(f_refs)} référents restaurés")
                    ecrire_meta("dernier_export", maintenant_iso())
                coffre["restaure"] = True
                st.cache_data.clear()
                if messages:
                    st.success("✅ " + " · ".join(messages))
                st.rerun()
            except Exception as exc:
                st.error(f"Chargement impossible : {exc}")
        if not pret:
            st.info("Charge les trois fichiers pour continuer.")
        st.stop()

    # ── Étape 2 : profil et mot de passe ─────────────────────────────────────
    st.subheader("2. Se connecter")
    st.caption(f"Fichier **{coffre['nom']}** chargé le {jolie_date(coffre['charge_le'], True)}.")

    gauche, _ = st.columns([1, 1])
    with gauche:
        with st.form("connexion"):
            nom = st.selectbox("Qui êtes-vous ?", list(PROFILS))
            mdp = st.text_input("Mot de passe", type="password")
            entrer = st.form_submit_button("Se connecter", type="primary", use_container_width=True)
        if entrer:
            if empreinte(mdp) == charger_empreinte(nom):
                st.session_state.utilisateur = nom
                st.session_state.role = PROFILS[nom]["role"]
                st.session_state.demo_active = True
                st.session_state.demo_etape = 0
                st.rerun()
            else:
                st.error("Mot de passe incorrect.")
        if st.button("📂 Charger d'autres fichiers"):
            coffre["contenu"] = None
            coffre["restaure"] = False
            coffre["clients"] = None
            st.cache_data.clear()
            st.rerun()
    st.stop()


if "utilisateur" not in st.session_state:
    ecran_connexion()

UTILISATEUR = st.session_state.utilisateur
ROLE = st.session_state.role
COULEUR_MOI = PROFILS[UTILISATEUR]["couleur"]

# Droits dérivés du rôle, plutôt que des tests dispersés sur le nom du rôle.
PEUT_TRAITER = ROLE in ("commercial", "admin")   # onglets appels, points, export
PEUT_PILOTER = ROLE in ("manager", "admin")      # onglet tableau de bord
PEUT_REINITIALISER = ROLE in ROLES_RESET         # exclusif au manager


# ─────────────────────────────────────────────────────────────────────────────
# CHARGEMENT DES DONNÉES
# ─────────────────────────────────────────────────────────────────────────────

if coffre.get("clients") is None:
    preparer_fichier()
clients = coffre["clients"]
col_code = coffre["col_code"]
adresses = coffre["adresses"]
if col_code is None:
    st.error("La feuille « Clients » doit contenir une colonne « Code client ».")
    st.stop()

suivi = charger_suivi()
suivi_adr = charger_suivi_adresses()

base = clients.merge(suivi, left_on=col_code, right_on="code_client", how="left")
for c in ["statut", "existe", "produits", "email_maj", "tel_maj", "note",
          "doublon_de", "rappel_date", "motif_sortie", "traite_par"]:
    base[c] = base[c].fillna("") if c in base.columns else ""
base["statut"] = base["statut"].replace("", "À appeler")
base["priorite"] = base["Type client"].map(priorite)

modifs_en_attente = nb_modifs_depuis_export()
rythme_global = calculer_rythme(suivi)
stats = {nom: stats_utilisateur(suivi, suivi_adr, nom) for nom in COMMERCIALES}

# Calculés une seule fois : la barre latérale et le bandeau les réutilisent.
PEUT_TRAITER_TMP = ROLE in ("commercial", "admin")
mes_rappels = rappels_dus(base, UTILISATEUR) if PEUT_TRAITER_TMP else pd.DataFrame()
tous_rappels = rappels_dus(base)
# Rappels programmés, passés comme à venir : matière du calendrier.
mon_calendrier = rappels_planifies(base, UTILISATEUR) if PEUT_TRAITER_TMP else pd.DataFrame()
calendrier_equipe = rappels_planifies(base)


# ─────────────────────────────────────────────────────────────────────────────
# BARRE LATÉRALE
# ─────────────────────────────────────────────────────────────────────────────

with st.sidebar:
    st.markdown(
        f"<div style='background:{COULEUR_MOI};color:#fff;border-radius:10px;"
        f"padding:12px 16px;font-weight:700;margin-bottom:12px'>"
        f"{UTILISATEUR}<br><span style='font-weight:400;font-size:0.82rem'>"
        f"{LIBELLES_ROLES.get(ROLE, ROLE)}</span></div>",
        unsafe_allow_html=True,
    )
    # Rafraîchissement : les données du suivi sont relues à chaque exécution,
    # ce bouton force donc l'affichage des saisies faites par les autres depuis
    # l'ouverture de la page.
    if st.button("🔄 Actualiser les données", use_container_width=True, type="primary",
                 help="Recharge le suivi pour voir en temps réel ce que les autres "
                      "ont enregistré depuis votre dernière action."):
        st.cache_data.clear()
        st.rerun()
    st.caption(f"Données à jour au {dt.datetime.now().strftime('%H:%M:%S')}")

    sc1, sc2 = st.columns(2)
    if sc1.button("🎓 Revoir la démo", use_container_width=True):
        st.session_state.demo_active = True
        st.session_state.demo_etape = 0
        st.rerun()
    if sc2.button("🚪 Déconnexion", use_container_width=True):
        for cle in ("utilisateur", "role", "idx", "idx_adr", "demo_active",
                    "demo_etape", "saut_code", "jour_rappel", "cal_decalage",
                    "grille_jour_detail", "bande_jour_detail"):
            st.session_state.pop(cle, None)
        st.rerun()

    if PEUT_TRAITER:
        st.divider()
        st.header("Mes chiffres")
        moi = stats_utilisateur(suivi, suivi_adr, UTILISATEUR)
        st.metric("Fiches traitées aujourd'hui", formater_entier(moi["aujourdhui"]))
        st.metric("Fiches traitées au total", formater_entier(moi["fiches"]))
        st.metric("Points de livraison vérifiés", formater_entier(moi["points"]))
        nb_dus = len(mes_rappels)
        if nb_dus:
            st.error(f"⏰ **{nb_dus} rappel(s) à passer aujourd'hui** — voir le bandeau en haut de page.")
        elif moi["a_rappeler"]:
            st.info(f"⏰ {moi['a_rappeler']} rappel(s) programmé(s) à votre nom, aucun dû aujourd'hui.")
        if modifs_en_attente > 0:
            st.error(f"⚠️ {modifs_en_attente} modification(s) non exportée(s) — onglet Export.")
        else:
            st.success("✅ Travail exporté : rien en attente.")

    st.divider()
    st.header("Avancement de l'équipe")
    total = len(base)
    faits = int(base["statut"].isin(STATUTS_TERMINES).sum())
    reste = total - faits
    st.metric("Clients au total", formater_entier(total))
    st.metric("Traités", formater_entier(faits), f"{(100 * faits / total):.1f} %" if total else "—")
    st.metric("Restants", formater_entier(reste))

    if PEUT_PILOTER:
        st.divider()
        st.header("Par commerciale")
        for nom in COMMERCIALES:
            s = stats[nom]
            st.markdown(
                f"<span class='pill' style='background:{PROFILS[nom]['couleur']}'>{nom}</span>",
                unsafe_allow_html=True,
            )
            m1, m2 = st.columns(2)
            m1.metric("Aujourd'hui", formater_entier(s["aujourdhui"]))
            m2.metric("Total", formater_entier(s["fiches"]))

    if rythme_global:
        st.divider()
        st.caption("Projection au rythme observé")
        st.metric("Fiches / jour actif (équipe)", f"{rythme_global:.0f}")
        st.metric("Fin estimée",
                  date_apres_jours_ouvres(dt.date.today(), round(reste / rythme_global)).strftime("%d/%m/%Y"))

    st.divider()
    st.caption(f"Objectif pour le {DEADLINE.strftime('%d/%m/%Y')}")
    jo = jours_ouvres(dt.date.today(), DEADLINE)
    if jo <= 0:
        st.error("Échéance atteinte ou dépassée.")
    else:
        objectif = -(-reste // jo)
        st.metric("Jours ouvrés restants", str(jo))
        st.metric("À traiter / jour (équipe)", str(objectif))
        if rythme_global:
            if rythme_global >= objectif:
                st.success(f"Rythme équipe ({rythme_global:.0f}/j) au-dessus de l'objectif.")
            else:
                st.error(f"Rythme équipe ({rythme_global:.0f}/j) sous l'objectif "
                         f"de ~{objectif - rythme_global:.0f}/j.")

    # Filtres : uniquement pour les profils qui ont une file d'appel.
    if PEUT_TRAITER:
        st.divider()
        st.header("🔎 Accès direct")
        code_direct = st.text_input("Code client exact")

        st.divider()
        st.header("Filtres")
        f_traitement = st.multiselect(
            "Traitement",
            ["Non traité"] + [f"Traité par {n}" for n in TRAITANTS],
            default=["Non traité"],
            help="Le filtre qui évite les doublons : « Non traité » ne montre que les fiches "
                 "que personne n'a encore prises.",
        )
        f_masquer_ouvertes = st.checkbox(
            "Masquer les fiches ouvertes par quelqu'un d'autre", value=True,
            help=f"Une fiche consultée par une autre personne depuis moins de {VERROU_MINUTES} "
                 "minutes est retirée de votre file.",
        )
        f_type = st.multiselect("Type de client", sorted(base["Type client"].unique()))
        f_statut = st.multiselect("Statut d'appel", STATUTS, default=["À appeler", "À rappeler"])
        f_acompl = st.checkbox("Uniquement « À compléter » non vide", value=False)
        recherche = st.text_input("Recherche (nom, code, ville)")
        tri_priorite = st.checkbox("Trier par priorité (pros d'abord)", value=True)
    else:
        code_direct, f_traitement, f_masquer_ouvertes = "", [], False
        f_type, f_statut, f_acompl, recherche, tri_priorite = [], [], False, "", True

    st.divider()
    with st.expander("⚙️ Réinitialisation (zone sensible)"):
        if not PEUT_REINITIALISER:
            st.caption("Réservé au profil manager, y compris pour l'administrateur : "
                       "un effacement total ne doit jamais pouvoir partir d'un poste de saisie.")
        else:
            st.caption("Efface tout le suivi : appels et référents. Irréversible. Exporter avant.")
            confirme = st.checkbox("Je comprends que tout sera effacé")
            mot_confirm = st.text_input("Écris RESET pour confirmer", value="")
            if st.button("🗑️ Tout réinitialiser",
                         disabled=not (confirme and mot_confirm.strip().upper() == "RESET")):
                reinitialiser_tout(UTILISATEUR)
                for cle in ("idx", "idx_adr"):
                    st.session_state.pop(cle, None)
                st.cache_data.clear()
                st.success("Suivi réinitialisé.")
                st.rerun()


# Démo proposée à l'ouverture de session, ou rappelée depuis la barre latérale.
if st.session_state.get("demo_active"):
    afficher_demo(ROLE)


# ─────────────────────────────────────────────────────────────────────────────
# NOTIFICATION DES RAPPELS DUS
# Affichée en haut de page, avant les onglets : c'est le seul endroit qu'on ne
# peut pas manquer, quel que soit l'onglet ouvert.
# ─────────────────────────────────────────────────────────────────────────────

aujourdhui = dt.date.today()

def _detail_rappels(df: pd.DataFrame, avec_bouton: bool) -> None:
    """Liste des clients à rappeler, avec accès direct à leur fiche."""
    for _, r in df.iterrows():
        retard = (aujourdhui - r["date_rappel"]).days
        quand = "aujourd'hui" if retard == 0 else f"en retard de {retard} jour(s)"
        c1, c2 = st.columns([5, 1])
        c1.markdown(
            f"**{r['Raison sociale / Nom']}** — {r.get('Ville', '')} · "
            f"☎️ {r.get('Téléphone 1', '') or '—'} · rappel {quand}"
            + (f" · noté par {r['traite_par']}" if r.get("traite_par") else "")
            + (f"<br><span style='color:#5a6b62;font-size:0.85rem'>{r['note']}</span>"
               if r.get("note") else ""),
            unsafe_allow_html=True,
        )
        if avec_bouton and c2.button("Ouvrir", key=f"saut_{r[col_code]}", use_container_width=True):
            st.session_state.saut_code = str(r[col_code])
            st.rerun()

if PEUT_TRAITER and not mes_rappels.empty:
    en_retard = int((mes_rappels["date_rappel"] < aujourdhui).sum())
    complement = f", dont {en_retard} en retard" if en_retard else ""
    st.error(f"### ⏰ {len(mes_rappels)} client(s) à rappeler aujourd'hui{complement}")
    with st.expander("Voir la liste et ouvrir les fiches", expanded=True):
        _detail_rappels(mes_rappels, avec_bouton=True)
elif PEUT_PILOTER and not tous_rappels.empty:
    en_retard = int((tous_rappels["date_rappel"] < aujourdhui).sum())
    complement = f", dont {en_retard} en retard" if en_retard else ""
    st.warning(f"### ⏰ {len(tous_rappels)} rappel(s) dû(s) dans l'équipe{complement}")
    with st.expander("Voir la liste"):
        _detail_rappels(tous_rappels, avec_bouton=False)

# Les rappels des autres, pour information, quand on a déjà traité les siens.
if PEUT_TRAITER and mes_rappels.empty and not tous_rappels.empty:
    st.info(f"⏰ {len(tous_rappels)} rappel(s) dû(s) dans l'équipe, aucun à votre nom.")


# ─────────────────────────────────────────────────────────────────────────────
# FILE D'APPEL
# ─────────────────────────────────────────────────────────────────────────────

file_appel = base.copy()

# Un saut depuis le bandeau de rappels court-circuite les filtres, comme
# l'accès direct par code.
code_saut = st.session_state.get("saut_code", "")
code_cible = str(code_direct).strip() or code_saut
acces_direct = bool(code_cible)

# Un jour choisi dans le calendrier remplace la file par les rappels de ce jour.
jour_choisi = st.session_state.get("jour_rappel", "")

if acces_direct:
    cible = str(code_cible).strip().upper()
    direct = base[base[col_code].astype(str).str.upper() == cible]
    if direct.empty:
        st.sidebar.error(f"Aucun client avec le code « {code_cible} ».")
        st.session_state.pop("saut_code", None)
        acces_direct = False
    else:
        file_appel = direct.reset_index(drop=True)
        st.session_state.idx = 0

if not acces_direct and jour_choisi and PEUT_TRAITER:
    source = mon_calendrier
    if jour_choisi == "retard":
        selection = source[source["date_rappel"] < dt.date.today()] if not source.empty else source
    else:
        cible_jour = dt.date.fromisoformat(jour_choisi)
        selection = source[source["date_rappel"] == cible_jour] if not source.empty else source
    if selection.empty:
        st.session_state.pop("jour_rappel", None)
        jour_choisi = ""
    else:
        file_appel = selection.drop(columns=["date_rappel"]).reset_index(drop=True)
        acces_direct = True   # les autres filtres ne s'appliquent pas

if not acces_direct and PEUT_TRAITER:
    if f_traitement:
        masques = []
        if "Non traité" in f_traitement:
            masques.append(file_appel["traite_par"].astype(str).str.strip() == "")
        for nom in TRAITANTS:
            if f"Traité par {nom}" in f_traitement:
                masques.append(file_appel["traite_par"] == nom)
        if masques:
            garde = masques[0]
            for m in masques[1:]:
                garde = garde | m
            file_appel = file_appel[garde]
    if f_masquer_ouvertes:
        occupees = fiches_ouvertes_par_les_autres("client", UTILISATEUR)
        if occupees:
            file_appel = file_appel[~file_appel[col_code].astype(str).isin(occupees)]
    if f_type:
        file_appel = file_appel[file_appel["Type client"].isin(f_type)]
    if f_statut:
        file_appel = file_appel[file_appel["statut"].isin(f_statut)]
    if f_acompl and "À compléter" in file_appel.columns:
        file_appel = file_appel[file_appel["À compléter"].astype(str).str.strip() != ""]
    if recherche:
        r = recherche.lower()
        file_appel = file_appel[
            file_appel[col_code].astype(str).str.lower().str.contains(r, na=False)
            | file_appel["Raison sociale / Nom"].str.lower().str.contains(r, na=False)
            | file_appel["Ville"].str.lower().str.contains(r, na=False)
        ]
    tri = ["priorite", "Raison sociale / Nom"] if tri_priorite else ["Raison sociale / Nom"]
    file_appel = file_appel.sort_values(tri).reset_index(drop=True)


# ─────────────────────────────────────────────────────────────────────────────
# ONGLETS SELON LE RÔLE
# ─────────────────────────────────────────────────────────────────────────────

# Les onglets visibles découlent des droits, pas d'un rôle en dur : le profil
# administrateur voit les quatre, les autres voient les leurs.
libelles_onglets = []
if PEUT_TRAITER:
    libelles_onglets += ["☎️  Appels clients", "📦  Points de livraison", "⬇️  Export"]
if PEUT_PILOTER:
    libelles_onglets += ["📊  Tableau de bord"]

onglets = st.tabs(libelles_onglets)
onglet_appel = onglet_adr = onglet_export = onglet_dash = None
curseur = 0
if PEUT_TRAITER:
    onglet_appel, onglet_adr, onglet_export = onglets[curseur:curseur + 3]
    curseur += 3
if PEUT_PILOTER:
    onglet_dash = onglets[curseur]


# ── ONGLETS DE TRAVAIL (commerciales et administrateur) ──────────────────────
if PEUT_TRAITER:
    with onglet_appel:
        # Calendrier de la semaine, replié par défaut pour ne pas repousser
        # la fiche hors de l'écran.
        with st.expander("🗓️ Mes rappels des sept prochains jours",
                         expanded=not st.session_state.get("jour_rappel")):
            bande_sept_jours(mon_calendrier, "bande")

        if st.session_state.get("jour_rappel"):
            choix = st.session_state.jour_rappel
            if choix == "retard":
                libelle = "rappels en retard"
                du_jour = (mon_calendrier[mon_calendrier["date_rappel"] < dt.date.today()]
                           if not mon_calendrier.empty else mon_calendrier)
            else:
                libelle = "rappels du " + jolie_date(choix)
                cible = dt.date.fromisoformat(choix)
                du_jour = (mon_calendrier[mon_calendrier["date_rappel"] == cible]
                           if not mon_calendrier.empty else mon_calendrier)

            st.info("File filtrée sur les " + libelle + " — " + str(len(file_appel)) + " fiche(s). "
                    "Les fiches ci-dessous s'enchaînent normalement avec « Enregistrer ».")
            # Qui appeler, à quel numéro, et ce qui avait été noté la fois d'avant.
            with st.expander("📋 La liste des clients à rappeler", expanded=True):
                tableau_rappels(du_jour, col_code)
            if st.button("↩︎ Revenir à ma file d'appel", key="retour_jour"):
                st.session_state.pop("jour_rappel", None)
                st.session_state.pop("idx", None)
                st.rerun()

        if st.session_state.get("saut_code"):
            if st.button("↩︎ Revenir à ma file d'appel"):
                st.session_state.pop("saut_code", None)
                st.session_state.pop("idx", None)
                st.rerun()

        if file_appel.empty:
            st.success("Aucun client dans la file avec ces filtres. 🎉")
            st.caption("Modifie les filtres à gauche, ou passe à l'onglet Export.")
        else:
            if "idx" not in st.session_state:
                derniere = lire_meta(f"derniere_fiche_{UTILISATEUR}", "")
                positions = file_appel.index[file_appel[col_code].astype(str) == derniere].tolist()
                st.session_state.idx = positions[0] if positions else 0
            st.session_state.idx = max(0, min(st.session_state.idx, len(file_appel) - 1))

            nav1, nav2, nav3 = st.columns([1, 2, 1])
            if nav1.button("⬅️ Précédent", use_container_width=True):
                st.session_state.idx = max(0, st.session_state.idx - 1)
                st.rerun()
            if nav3.button("Suivant ➡️", use_container_width=True):
                st.session_state.idx = min(len(file_appel) - 1, st.session_state.idx + 1)
                st.rerun()
            nav2.markdown(
                f"<div style='text-align:center;font-weight:600;color:{VERT}'>"
                f"Fiche {st.session_state.idx + 1} / {len(file_appel)}</div>",
                unsafe_allow_html=True,
            )

            ligne = file_appel.iloc[st.session_state.idx]
            code = str(ligne[col_code])
            marquer_presence("client", code, UTILISATEUR)

            # Bandeau si quelqu'un d'autre est sur la même fiche.
            for qui, minutes in autres_sur_la_fiche("client", code, UTILISATEUR):
                st.markdown(
                    f"<div class='bandeau-verrou'>⚠️ <b>{qui}</b> consulte cette fiche "
                    f"{'à l’instant' if minutes < 1 else f'depuis {minutes} min'}. "
                    f"Pour éviter d'appeler deux fois le même client, passez à la suivante.</div>",
                    unsafe_allow_html=True,
                )

            gauche, droite = st.columns([3, 2])

            with gauche:
                st.markdown(f"### {ligne['Raison sociale / Nom']}")
                st.markdown(
                    pastille_traitement(ligne.get("traite_par", ""))
                    + f"<span class='pill'>{ligne['Type client']}</span>"
                    + f"<span class='pill pill-orange'>{ligne.get('Catégorie', '')}</span>"
                    + f"<span style='color:#5a6b62'>Code {code}</span>",
                    unsafe_allow_html=True,
                )
                if ligne.get("maj_le"):
                    st.caption(f"Dernière mise à jour : {jolie_date(ligne['maj_le'], True)}")

                adresse = " ".join(x for x in [ligne.get("Adresse 1", ""), ligne.get("Adresse 2", ""),
                                               ligne.get("Adresse 3", "")] if x)
                email_sec = ligne.get("Email secondaire", "")
                st.markdown(
                    f"""<div class='fiche'>
                    📍 {adresse}<br>{ligne.get('Code postal', '')} {ligne.get('Ville', '')}<br><br>
                    ☎️ {ligne.get('Téléphone 1', '')} &nbsp; {ligne.get('Téléphone 2', '')} &nbsp; {ligne.get('Téléphone 3', '')}<br>
                    ✉️ {ligne.get('Email principal', '') or '<i>aucun e-mail</i>'}{(' · ' + email_sec) if email_sec else ''}<br>
                    🏢 SIREN : {ligne.get('SIREN', '') or '<i>—</i>'}
                    </div>""",
                    unsafe_allow_html=True,
                )
                if ligne.get("À compléter", ""):
                    st.warning(f"À compléter : {ligne['À compléter']}")

                if not adresses.empty:
                    liees = adresses[adresses["Code client mère"] == code]
                    if not liees.empty:
                        with st.expander(f"📦 {len(liees)} adresse(s) de livraison rattachée(s)"):
                            apercu = liees[["Code adresse", "Nom site", "Adresse 1", "Code postal", "Ville"]].merge(
                                suivi_adr[["code_adresse", "statut_adr", "traite_par"]]
                                if not suivi_adr.empty else
                                pd.DataFrame(columns=["code_adresse", "statut_adr", "traite_par"]),
                                left_on="Code adresse", right_on="code_adresse", how="left",
                            ).fillna("")
                            apercu["statut_adr"] = apercu["statut_adr"].replace("", "À vérifier")
                            apercu["traite_par"] = apercu["traite_par"].replace("", "Non vérifié")
                            st.dataframe(
                                apercu[["Code adresse", "Nom site", "Ville", "statut_adr", "traite_par"]]
                                .rename(columns={"statut_adr": "Statut", "traite_par": "Vérifié par"}),
                                hide_index=True, use_container_width=True,
                            )
                            st.caption("Complétez ces points dans l'onglet « Points de livraison », "
                                       "en recherchant par code client mère.")

            with droite:
                st.markdown("#### Résultat de l'appel")
                produits_init = [p for p in str(ligne.get("produits") or "").split("|") if p in PRODUITS]
                rappel_init = None
                if ligne.get("rappel_date"):
                    try:
                        rappel_init = pd.to_datetime(ligne["rappel_date"]).date()
                    except Exception:
                        rappel_init = None

                # Chaque widget porte une clé incluant le code client : sans cela,
                # l'état d'une fiche déborderait sur la suivante.
                with st.form(f"appel_{code}", clear_on_submit=False):
                    statut = st.selectbox(
                        "Statut", STATUTS,
                        index=STATUTS.index(ligne["statut"]) if ligne["statut"] in STATUTS else 0,
                        key=f"statut_{code}")
                    existe = st.radio(
                        "Client toujours actif ?", ["Oui", "Non", "Incertain"], horizontal=True,
                        index=["Oui", "Non", "Incertain"].index(ligne.get("existe"))
                        if ligne.get("existe") in ("Oui", "Non", "Incertain") else 0,
                        key=f"existe_{code}")
                    produits = st.multiselect("Produits achetés", PRODUITS, default=produits_init,
                                              key=f"produits_{code}")
                    email_maj = st.text_input("E-mail confirmé / corrigé",
                                              value=ligne.get("email_maj") or "", key=f"email_{code}")
                    tel_maj = st.text_input("Téléphone confirmé / corrigé",
                                            value=ligne.get("tel_maj") or "", key=f"tel_{code}")
                    doublon_de = st.text_input(
                        "Doublon du client n°", value=ligne.get("doublon_de") or "",
                        help="Si ce client est un doublon, indiquer le code à conserver.",
                        key=f"doublon_{code}")
                    motif_sortie = st.selectbox(
                        "Motif de sortie (si ancien client)", MOTIFS_SORTIE,
                        index=MOTIFS_SORTIE.index(ligne.get("motif_sortie"))
                        if ligne.get("motif_sortie") in MOTIFS_SORTIE else 0,
                        key=f"motif_{code}")
                    # Champ de date toujours actif : à l'intérieur d'un formulaire,
                    # une case à cocher ne prend effet qu'à l'enregistrement, elle ne
                    # peut donc pas conditionner l'accès à un autre champ.
                    # La plage est volontairement large pour n'imposer aucune limite.
                    rappel = st.date_input(
                        "Date de rappel (laisser vide s'il n'y en a pas)",
                        value=rappel_init,
                        min_value=dt.date(2020, 1, 1),
                        max_value=dt.date.today() + dt.timedelta(days=365 * 5),
                        format="DD/MM/YYYY",
                        key=f"rappel_{code}",
                        help="Saisissez la date au clavier (JJ/MM/AAAA) ou choisissez-la "
                             "dans le calendrier.")
                    # La suppression d'un rappel existant reste un geste explicite.
                    supprimer_rappel = False
                    if rappel_init:
                        supprimer_rappel = st.checkbox(
                            f"Supprimer le rappel du {rappel_init.strftime('%d/%m/%Y')}",
                            value=False, key=f"suppr_rappel_{code}")
                    note = st.text_area("Notes (commercial, vérifications…)",
                                        value=ligne.get("note") or "", height=90, key=f"note_{code}")
                    with st.expander("Effacer volontairement un champ"):
                        st.caption(
                            "Par défaut, un champ laissé vide ne remplace jamais une information "
                            "déjà enregistrée. Cochez ici ce que vous voulez réellement supprimer."
                        )
                        a_effacer = st.multiselect(
                            "Champs à vider", list(LIBELLES_CHAMPS), default=[],
                            key=f"effacer_{code}", label_visibility="collapsed")
                    valide = st.form_submit_button("💾 Enregistrer & passer au suivant",
                                                   use_container_width=True, type="primary")

                if valide:
                    erreurs = []
                    if email_maj.strip() and not re.match(r"^[^@\s]+@[^@\s]+\.[a-zA-Z]{2,}$", email_maj.strip()):
                        erreurs.append("l'adresse e-mail ne semble pas valide")
                    if statut == "Doublon" and not doublon_de.strip():
                        erreurs.append("indique le code du client à conserver")
                    if statut == "Ancien client (à sortir)" and motif_sortie == "—":
                        erreurs.append("indique le motif de sortie")
                    if statut == "À rappeler" and (not rappel or supprimer_rappel):
                        erreurs.append("indique une date de rappel")

                    if erreurs:
                        st.error("Avant d'enregistrer : " + ", ".join(erreurs) + ".")
                    else:
                        # Seuls les champs explicitement désignés peuvent être vidés.
                        effacements = [LIBELLES_CHAMPS[lib] for lib in a_effacer]
                        if supprimer_rappel:
                            effacements.append("rappel_date")
                        enregistrer(
                            code, UTILISATEUR, effacements=effacements,
                            statut=statut, existe=existe,
                            produits="|".join(produits),
                            email_maj=email_maj.strip(), tel_maj=tel_maj.strip(),
                            doublon_de=doublon_de.strip(), note=note.strip(),
                            motif_sortie="" if motif_sortie == "—" else motif_sortie,
                            rappel_date="" if supprimer_rappel else (rappel.isoformat() if rappel else ""),
                        )
                        sauvegarde_auto()
                        st.session_state.idx = min(len(file_appel) - 1, st.session_state.idx + 1)
                        st.rerun()

                # ── Réattribution manuelle, réservée à l'administrateur ──────
                # Permet de corriger une attribution : rendre une fiche à la
                # file (« Non traité »), la basculer d'une personne à l'autre,
                # ou forcer un statut sans repasser par le formulaire d'appel.
                if ROLE == "admin":
                    st.divider()
                    with st.expander("🛠️ Administration de la fiche", expanded=False):
                        st.caption(
                            "Réservé à l'administrateur. Modifie l'attribution et le statut "
                            "sans toucher aux autres informations de la fiche. "
                            "Chaque intervention est inscrite au journal."
                        )
                        actuel = str(ligne.get("traite_par") or "")
                        options_qui = ["Non traité"] + TRAITANTS
                        with st.form(f"admin_fiche_{code}"):
                            nouveau_qui = st.selectbox(
                                "Attribuer la fiche à", options_qui,
                                index=options_qui.index(actuel) if actuel in options_qui else 0,
                                key=f"admin_qui_{code}",
                            )
                            nouveau_statut = st.selectbox(
                                "Forcer le statut", STATUTS,
                                index=STATUTS.index(ligne["statut"]) if ligne["statut"] in STATUTS else 0,
                                key=f"admin_statut_{code}",
                            )
                            appliquer = st.form_submit_button(
                                "Appliquer la modification", use_container_width=True)
                        if appliquer:
                            traitant = "" if nouveau_qui == "Non traité" else nouveau_qui
                            if traitant == actuel and nouveau_statut == ligne["statut"]:
                                st.info("Aucun changement à appliquer.")
                            else:
                                reattribuer(
                                    code, traitant, nouveau_statut, UTILISATEUR,
                                    {"traite_par": actuel, "statut": ligne["statut"]},
                                )
                                sauvegarde_auto()
                                st.success(
                                    f"Fiche {code} : attribuée à "
                                    f"{traitant or 'personne'}, statut « {nouveau_statut} »."
                                )
                                st.rerun()


    # ── ONGLET POINTS DE LIVRAISON ───────────────────────────────────────────
    with onglet_adr:
        if adresses.empty:
            st.info("Le fichier ne contient pas de feuille « Adresses livraison ».")
        else:
            st.subheader("Vérification des points de livraison")
            st.caption("Prolongement de l'appel : pour chaque site livré, qui est le référent sur place ?")

            adr = adresses.merge(suivi_adr, left_on="Code adresse", right_on="code_adresse", how="left")
            for c in ["referent", "tel_site", "statut_adr", "note_adr", "traite_par"]:
                adr[c] = adr[c].fillna("") if c in adr.columns else ""
            adr["statut_adr"] = adr["statut_adr"].replace("", "À vérifier")

            m1, m2, m3 = st.columns(3)
            m1.metric("Points de livraison", formater_entier(len(adr)))
            m2.metric("Vérifiés", formater_entier(int((adr["statut_adr"] == "Vérifié ✅").sum())))
            m3.metric("Restants", formater_entier(int((adr["statut_adr"] != "Vérifié ✅").sum())))

            rech1, rech2 = st.columns(2)
            q_adr = rech1.text_input("🔎 Code adresse exact (ex. 12771L56)")
            q_mere = rech2.text_input("🔎 ou Code client mère (montre tous ses points)")

            vue = adr.copy()
            if q_adr.strip():
                vue = vue[vue["Code adresse"].astype(str).str.upper() == q_adr.strip().upper()]
            elif q_mere.strip():
                vue = vue[vue["Code client mère"].astype(str).str.upper() == q_mere.strip().upper()]
            else:
                fa1, fa2 = st.columns(2)
                filtre_adr = fa1.multiselect("Statut", STATUTS_ADRESSE, default=["À vérifier"])
                filtre_qui = fa2.multiselect(
                    "Traitement", ["Non vérifié"] + [f"Vérifié par {n}" for n in TRAITANTS],
                    default=[])
                if filtre_adr:
                    vue = vue[vue["statut_adr"].isin(filtre_adr)]
                if filtre_qui:
                    masques = []
                    if "Non vérifié" in filtre_qui:
                        masques.append(vue["traite_par"].astype(str).str.strip() == "")
                    for nom in TRAITANTS:
                        if f"Vérifié par {nom}" in filtre_qui:
                            masques.append(vue["traite_par"] == nom)
                    if masques:
                        garde = masques[0]
                        for m in masques[1:]:
                            garde = garde | m
                        vue = vue[garde]
            vue = vue.reset_index(drop=True)

            if vue.empty:
                st.success("Aucun point de livraison à afficher avec ce filtre.")
            else:
                if "idx_adr" not in st.session_state:
                    st.session_state.idx_adr = 0
                st.session_state.idx_adr = max(0, min(st.session_state.idx_adr, len(vue) - 1))

                n1, n2, n3 = st.columns([1, 2, 1])
                if n1.button("⬅️ Précédent", key="adr_prev", use_container_width=True):
                    st.session_state.idx_adr = max(0, st.session_state.idx_adr - 1)
                    st.rerun()
                if n3.button("Suivant ➡️", key="adr_next", use_container_width=True):
                    st.session_state.idx_adr = min(len(vue) - 1, st.session_state.idx_adr + 1)
                    st.rerun()
                n2.markdown(
                    f"<div style='text-align:center;font-weight:600;color:{VERT}'>"
                    f"Point {st.session_state.idx_adr + 1} / {len(vue)}</div>",
                    unsafe_allow_html=True,
                )

                point = vue.iloc[st.session_state.idx_adr]
                code_adr = str(point["Code adresse"])
                marquer_presence("adresse", code_adr, UTILISATEUR)

                for qui, minutes in autres_sur_la_fiche("adresse", code_adr, UTILISATEUR):
                    st.markdown(
                        f"<div class='bandeau-verrou'>⚠️ <b>{qui}</b> consulte ce point de livraison "
                        f"{'à l’instant' if minutes < 1 else f'depuis {minutes} min'}.</div>",
                        unsafe_allow_html=True,
                    )

                adresse_txt = " ".join(x for x in [point.get("Adresse 1", ""), point.get("Adresse 2", ""),
                                                   point.get("Adresse 3", "")] if x)
                g, d = st.columns([3, 2])
                with g:
                    st.markdown(f"### {point.get('Nom site', '') or 'Point de livraison'}")
                    st.markdown(
                        pastille_traitement(point.get("traite_par", ""))
                        + f"<span class='pill'>Adresse {code_adr}</span>"
                        + f"<span style='color:#5a6b62'>Client mère : {point.get('Code client mère', '')}</span>",
                        unsafe_allow_html=True,
                    )
                    st.markdown(
                        f"""<div class='fiche'>
                        📍 {adresse_txt}<br>{point.get('Code postal', '')} {point.get('Ville', '')}<br><br>
                        ☎️ {point.get('Téléphone', '') or '<i>aucun téléphone</i>'}
                        </div>""",
                        unsafe_allow_html=True,
                    )
                with d:
                    st.markdown("#### Référent du site")
                    with st.form(f"adr_form_{code_adr}"):
                        referent = st.text_input("Nom du référent sur place",
                                                 value=point.get("referent") or "", key=f"referent_{code_adr}")
                        tel_site = st.text_input("Téléphone du site / référent",
                                                 value=point.get("tel_site") or "", key=f"tel_site_{code_adr}")
                        statut_adr = st.selectbox(
                            "Statut", STATUTS_ADRESSE,
                            index=STATUTS_ADRESSE.index(point["statut_adr"])
                            if point["statut_adr"] in STATUTS_ADRESSE else 0,
                            key=f"statut_adr_{code_adr}")
                        note_adr = st.text_area("Note", value=point.get("note_adr") or "",
                                                height=80, key=f"note_adr_{code_adr}")
                        with st.expander("Effacer volontairement un champ"):
                            st.caption("Un champ laissé vide ne remplace jamais une information "
                                       "déjà enregistrée.")
                            a_effacer_adr = st.multiselect(
                                "Champs à vider",
                                ["Référent sur place", "Téléphone du site", "Note"],
                                default=[], key=f"effacer_adr_{code_adr}",
                                label_visibility="collapsed")
                        valide_adr = st.form_submit_button("💾 Enregistrer & suivant",
                                                           use_container_width=True, type="primary")
                    if valide_adr:
                        corresp_adr = {"Référent sur place": "referent",
                                       "Téléphone du site": "tel_site", "Note": "note_adr"}
                        enregistrer_adresse(
                            code_adr, UTILISATEUR,
                            effacements=[corresp_adr[x] for x in a_effacer_adr],
                            referent=referent.strip(), tel_site=tel_site.strip(),
                            statut_adr=statut_adr, note_adr=note_adr.strip(),
                        )
                        sauvegarde_auto()
                        st.session_state.idx_adr = min(len(vue) - 1, st.session_state.idx_adr + 1)
                        st.rerun()


    # ── ONGLET EXPORT ────────────────────────────────────────────────────────
    with onglet_export:
        st.subheader("Export du suivi")
        if modifs_en_attente > 0:
            st.warning(
                f"🔔 {modifs_en_attente} modification(s) à exporter. "
                "Télécharge les deux fichiers avant de fermer l'onglet."
            )
        else:
            st.success("✅ Tout est exporté : rien en attente.")
        st.caption(
            "Ces fichiers sont la sauvegarde de la campagne et servent à rouvrir une session "
            "le lendemain. La donnée de référence reste Logimatique."
        )

        st.markdown("##### Suivi des appels clients")
        if suivi.empty:
            st.info("Aucun appel enregistré pour le moment.")
        else:
            export = preparer_export(suivi, base, col_code)
            e1, e2 = st.columns(2)
            e1.download_button(
                "⬇️ Suivi clients — CSV (Excel FR)",
                export.to_csv(index=False, sep=";").encode("utf-8-sig"),
                file_name=f"suivi_appels_hympyr_{dt.date.today():%Y%m%d}.csv",
                mime="text/csv", use_container_width=True)
            e2.download_button(
                "⬇️ Suivi clients — Excel (.xlsx)",
                vers_excel(export, "Suivi appels"),
                file_name=f"suivi_appels_hympyr_{dt.date.today():%Y%m%d}.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                use_container_width=True)
            with st.expander("Aperçu de l'export clients"):
                st.dataframe(export, hide_index=True, use_container_width=True)

        if not adresses.empty:
            st.divider()
            st.markdown("##### Points de livraison et référents")
            export_adr = preparer_export_adresses(suivi_adr, adresses)
            if export_adr.empty:
                st.info("Aucun point de livraison à exporter.")
            else:
                a1, a2 = st.columns(2)
                a1.download_button(
                    "⬇️ Points de livraison — CSV (Excel FR)",
                    export_adr.to_csv(index=False, sep=";").encode("utf-8-sig"),
                    file_name=f"points_livraison_hympyr_{dt.date.today():%Y%m%d}.csv",
                    mime="text/csv", use_container_width=True)
                a2.download_button(
                    "⬇️ Points de livraison — Excel (.xlsx)",
                    vers_excel(export_adr, "Points de livraison"),
                    file_name=f"points_livraison_hympyr_{dt.date.today():%Y%m%d}.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    use_container_width=True)
                with st.expander("Aperçu de l'export points de livraison"):
                    st.dataframe(export_adr, hide_index=True, use_container_width=True)

        if modifs_en_attente > 0:
            st.divider()
            st.caption("Une fois les deux fichiers téléchargés, confirme pour repasser au vert :")
            if st.button("✅ J'ai bien téléchargé mes sauvegardes", type="primary"):
                ecrire_meta("dernier_export", maintenant_iso())
                st.rerun()


# ── ONGLET TABLEAU DE BORD (manager et administrateur) ───────────────────────
if PEUT_PILOTER:
    with onglet_dash:
        st.subheader("🎯 Objectif pour tenir l'échéance")
        st.caption(f"Échéance : émission de la facturation électronique au "
                   f"{DEADLINE.strftime('%d/%m/%Y')} pour les PME.")

        perimetre = st.radio(
            "Périmètre à boucler",
            ["Tous les clients restants", "Uniquement les pros (conformité)"],
            horizontal=True,
        )
        masque = (base["Type client"].astype(str).str.startswith("Pro")
                  if perimetre.startswith("Uniquement") else pd.Series(True, index=base.index))
        restant_perimetre = int((masque & ~base["statut"].isin(STATUTS_TERMINES)).sum())
        jo = jours_ouvres(dt.date.today(), DEADLINE)

        o1, o2, o3, o4 = st.columns(4)
        o1.metric("Restant sur ce périmètre", formater_entier(restant_perimetre))
        o2.metric("Jours ouvrés d'ici l'échéance", str(jo))

        if jo > 0:
            objectif = -(-restant_perimetre // jo)
            o3.metric("À traiter / jour (équipe)", str(objectif))
            if rythme_global:
                o4.metric("Rythme équipe / jour", f"{rythme_global:.0f}",
                          delta=f"{rythme_global - objectif:+.0f} vs objectif")
                if rythme_global >= objectif:
                    st.success(f"✅ Au rythme actuel ({rythme_global:.0f}/jour), "
                               "l'échéance est tenable sur ce périmètre.")
                else:
                    fin = date_apres_jours_ouvres(dt.date.today(), round(restant_perimetre / rythme_global))
                    st.error(f"⚠️ Au rythme actuel ({rythme_global:.0f}/jour), fin estimée vers le "
                             f"{fin.strftime('%d/%m/%Y')}, soit après l'échéance. "
                             f"Il faut viser {objectif}/jour, ou renforcer l'équipe.")
            else:
                o4.metric("Rythme équipe / jour", "—")
                st.info("Le rythme s'affichera après les premiers appels enregistrés.")
        else:
            o3.metric("À traiter / jour", "—")
            st.error("L'échéance est atteinte ou dépassée.")

        # ── Performance par commerciale ──────────────────────────────────────
        st.divider()
        st.subheader("Performance des commerciales")
        st.caption("Alimenté directement par la saisie de Chloé et de Patricia, sans ressaisie.")

        colonnes = st.columns(len(COMMERCIALES))
        for col, nom in zip(colonnes, COMMERCIALES):
            s = stats[nom]
            with col:
                st.markdown(
                    f"<span class='pill' style='background:{PROFILS[nom]['couleur']}'>{nom}</span>",
                    unsafe_allow_html=True)
                st.metric("Fiches traitées aujourd'hui", formater_entier(s["aujourdhui"]))
                st.metric("Fiches traitées au total", formater_entier(s["terminees"]),
                          help="Statuts Fait, Doublon ou Ancien client.")
                st.metric("Rythme / jour actif", f"{s['rythme']:.0f}" if s["rythme"] else "—")
                st.metric("Points de livraison vérifiés", formater_entier(s["points"]),
                          delta=f"+{s['points_aujourdhui']} aujourd'hui" if s["points_aujourdhui"] else None)
                st.metric("Rappels en cours", formater_entier(s["a_rappeler"]))

        tableau = pd.DataFrame([
            {
                "Commerciale": nom,
                "Fiches traitées": stats[nom]["fiches"],
                "Dont terminées": stats[nom]["terminees"],
                "Aujourd'hui": stats[nom]["aujourdhui"],
                "Rappels en cours": stats[nom]["a_rappeler"],
                "Points de livraison": stats[nom]["points"],
                "Rythme / jour actif": round(stats[nom]["rythme"], 1) if stats[nom]["rythme"] else 0,
            }
            for nom in COMMERCIALES
        ])
        st.dataframe(tableau, hide_index=True, use_container_width=True)

        non_traitees = int((base["traite_par"].astype(str).str.strip() == "").sum())
        st.caption(f"{formater_entier(non_traitees)} fiche(s) ne sont encore attribuées à personne.")

        # L'activité des profils administrateurs est comptée dans l'avancement
        # global, mais signalée à part : le suivi des commerciales reste lisible.
        admins = [n for n, prof in PROFILS.items() if prof["role"] == "admin"]
        for nom in admins:
            s_admin = stats_utilisateur(suivi, suivi_adr, nom)
            if s_admin["fiches"] or s_admin["points"]:
                st.caption(
                    f"Hors périmètre commercial : {formater_entier(s_admin['fiches'])} fiche(s) "
                    f"et {formater_entier(s_admin['points'])} point(s) de livraison traités par "
                    f"{nom} (administrateur)."
                )

        # Activité quotidienne comparée
        if not suivi.empty and (suivi["traite_par"] != "").any():
            activite = suivi[suivi["traite_par"].isin(COMMERCIALES)].copy()
            if not activite.empty:
                activite["jour"] = pd.to_datetime(activite["maj_le"], errors="coerce").dt.date
                pivot = (activite.dropna(subset=["jour"])
                         .pivot_table(index="jour", columns="traite_par",
                                      values="code_client", aggfunc="count")
                         .fillna(0))
                if not pivot.empty:
                    st.markdown("##### Fiches traitées par jour")
                    st.bar_chart(pivot)

        # ── Avancement global ────────────────────────────────────────────────
        st.divider()
        st.subheader("Avancement de la campagne")
        c1, c2, c3, c4 = st.columns(4)
        c1.metric("Total clients", formater_entier(len(base)))
        c2.metric("Traités", formater_entier(int(base["statut"].isin(STATUTS_TERMINES).sum())))
        c3.metric("À rappeler", formater_entier(int((base["statut"] == "À rappeler").sum())))
        c4.metric("Doublons repérés", formater_entier(int((base["statut"] == "Doublon").sum())))

        g1, g2 = st.columns(2)
        with g1:
            st.markdown("##### Répartition par statut")
            st.bar_chart(base["statut"].value_counts())
        with g2:
            st.markdown("##### Répartition par type de client")
            st.bar_chart(base["Type client"].value_counts())

        if not suivi.empty and suivi["produits"].str.len().gt(0).any():
            eclate = suivi["produits"].str.split("|").explode()
            eclate = eclate[eclate.isin(PRODUITS)]
            if not eclate.empty:
                st.markdown("##### Produits achetés (déclarés en appel)")
                st.bar_chart(eclate.value_counts())

        # ── Rappels du jour ou en retard ─────────────────────────────────────
        rappels = base[(base["statut"] == "À rappeler")
                       & (base["rappel_date"].astype(str).str.len() > 0)].copy()
        if not rappels.empty:
            rappels["date"] = pd.to_datetime(rappels["rappel_date"], errors="coerce").dt.date
            a_faire = rappels[rappels["date"].notna() & (rappels["date"] <= dt.date.today())]
            if not a_faire.empty:
                st.divider()
                st.markdown(f"##### ⏰ {len(a_faire)} rappel(s) à passer aujourd'hui ou en retard")
                st.dataframe(
                    a_faire[[col_code, "Raison sociale / Nom", "Ville", "Téléphone 1",
                             "traite_par", "rappel_date", "note"]]
                    .rename(columns={col_code: "Code client", "traite_par": "Traité par",
                                     "rappel_date": "À rappeler le", "note": "Notes"}),
                    hide_index=True, use_container_width=True,
                )

        # ── Calendrier des rappels ──────────────────────────────────────────
        st.divider()
        st.subheader("🗓️ Calendrier des rappels")
        st.caption("Alimenté par les dates que les commerciales inscrivent dans les fiches. "
                   "C'est la charge à venir, que les autres écrans ne montrent pas.")
        grille_mensuelle(calendrier_equipe, "grille", col_code)

        # ── Journal des interventions manuelles ─────────────────────────────
        journal = charger_journal()
        if not journal.empty:
            st.divider()
            with st.expander(f"🛠️ Journal des interventions manuelles ({len(journal)})"):
                st.caption("Réattributions et forçages de statut effectués par un administrateur.")
                affichage = journal.copy()
                affichage["horodatage"] = affichage["horodatage"].map(lambda v: jolie_date(v, True))
                st.dataframe(
                    affichage.rename(columns={
                        "horodatage": "Quand", "auteur": "Par qui",
                        "action": "Action", "cible": "Fiche", "detail": "Détail"}),
                    hide_index=True, use_container_width=True,
                )

        st.divider()
        if ROLE == "manager":
            st.caption(
                "Profil manager : lecture seule. Les exports sont réalisés par les commerciales "
                "depuis leur onglet « Export »."
            )
        else:
            st.caption(
                "Profil administrateur : accès complet aux onglets de travail et de pilotage. "
                "La réinitialisation reste réservée au profil manager."
            )
